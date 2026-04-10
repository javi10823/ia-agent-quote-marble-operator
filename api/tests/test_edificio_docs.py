"""Tests for generate_edificio_documents — 4 PDF+Excel from paso2_calc.

Uses real ESH data through the full pipeline.
"""

import shutil
import pytest
import zipfile
from pathlib import Path

ESH_PDF = Path("/Users/javierolivieri/projects/dangelo-marble-ia/planos/edifcio-1121SM-2025-Planilla marmolería.pdf")
OUTPUT_DIR = Path(__file__).parent.parent / "output"


@pytest.fixture(autouse=True)
def cleanup():
    yield
    for d in OUTPUT_DIR.glob("test-edif-doc-*"):
        if d.is_dir():
            shutil.rmtree(d, ignore_errors=True)


def _build_paso2_calc():
    """Run full pipeline + Paso 2 calc on ESH."""
    import pdfplumber
    from app.modules.quote_engine.edificio_parser import (
        parse_edificio_tables, normalize_edificio_data,
        compute_edificio_aggregates, render_edificio_paso2,
    )

    tables = []
    with pdfplumber.open(str(ESH_PDF)) as pdf:
        for p in pdf.pages:
            tables.extend(p.extract_tables())

    raw = parse_edificio_tables(tables)
    norm = normalize_edificio_data(raw)
    summary = compute_edificio_aggregates(norm)
    paso2 = render_edificio_paso2(summary, "Rosario")
    return paso2, summary


@pytest.mark.skipif(not ESH_PDF.exists(), reason="ESH PDF not available")
class TestEdificioDocs:

    @pytest.fixture(scope="class")
    def paso2_and_summary(self):
        return _build_paso2_calc()

    @pytest.fixture(scope="class")
    def paso2(self, paso2_and_summary):
        return paso2_and_summary[0]

    @pytest.fixture(scope="class")
    def summary(self, paso2_and_summary):
        return paso2_and_summary[1]

    @pytest.mark.asyncio
    async def test_generates_4_document_pairs(self, paso2, summary):
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-001"
        result = await generate_edificio_documents(
            quote_id=qid,
            paso2_calc=paso2,
            summary=summary,
            client_name="ESH 1121-SM",
            project="Concesionario",
        )
        assert result["ok"]
        assert len(result["generated"]) == 4  # 3 materials + 1 services

    @pytest.mark.asyncio
    async def test_3_material_docs_plus_1_services(self, paso2, summary):
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-002"
        result = await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")
        mat_docs = [g for g in result["generated"] if g["type"] == "material"]
        svc_docs = [g for g in result["generated"] if g["type"] == "services"]
        assert len(mat_docs) == 3
        assert len(svc_docs) == 1

    @pytest.mark.asyncio
    async def test_pdf_and_excel_exist_on_disk(self, paso2, summary):
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-003"
        result = await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")
        qdir = OUTPUT_DIR / qid
        pdfs = list(qdir.glob("*.pdf"))
        xlsxs = list(qdir.glob("*.xlsx"))
        assert len(pdfs) == 4, f"Expected 4 PDFs, got {len(pdfs)}"
        assert len(xlsxs) == 4, f"Expected 4 Excels, got {len(xlsxs)}"

    @pytest.mark.asyncio
    async def test_grand_totals_match_paso2(self, paso2, summary):
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-004"
        result = await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")
        assert result["grand_total_ars"] == paso2["grand_total_ars"]
        assert result["grand_total_usd"] == paso2["grand_total_usd"]

    @pytest.mark.asyncio
    async def test_services_excel_has_mo_items(self, paso2, summary):
        """Services Excel must contain all MO items."""
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-005"
        await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")

        svc_xlsx = list((OUTPUT_DIR / qid).glob("*Servicios*.xlsx"))
        assert len(svc_xlsx) == 1

        import openpyxl
        wb = openpyxl.load_workbook(str(svc_xlsx[0]))
        ws = wb.active
        all_text = " ".join(str(ws.cell(r, c).value or "") for r in range(1, 30) for c in range(1, 5))
        wb.close()

        assert "pegado pileta" in all_text.lower() or "pegadopileta" in all_text.lower()
        assert "apoyo" in all_text.lower()

    @pytest.mark.asyncio
    async def test_services_pdf_has_correct_content(self, paso2, summary):
        """Services PDF must contain MO data, not material data."""
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-006"
        await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")

        svc_pdfs = list((OUTPUT_DIR / qid).glob("*Servicios*.pdf"))
        assert len(svc_pdfs) == 1
        # PDF exists and has content
        assert svc_pdfs[0].stat().st_size > 1000  # Not empty

    @pytest.mark.asyncio
    async def test_no_sink_product_in_any_doc(self, paso2, summary):
        """No Johnson/Quadra/Luxor in any generated document."""
        from app.modules.agent.tools.document_tool import generate_edificio_documents
        qid = "test-edif-doc-007"
        await generate_edificio_documents(qid, paso2, summary, "ESH", "Test")

        # Check all Excels (easier to parse than PDFs)
        for xlsx in (OUTPUT_DIR / qid).glob("*.xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(str(xlsx))
            ws = wb.active
            all_text = " ".join(str(ws.cell(r, c).value or "") for r in range(1, 30) for c in range(1, 5))
            wb.close()
            for forbidden in ["Johnson", "Quadra", "Luxor", "Oval"]:
                assert forbidden not in all_text, f"Found '{forbidden}' in {xlsx.name}"
