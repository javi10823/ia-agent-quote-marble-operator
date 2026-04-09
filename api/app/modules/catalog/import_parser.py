"""Parse Dux export files (.xls/.xlsx/.csv) and classify items by catalog.

Supports three Dux export formats:
1. Materials (USD): has Costo + Porc. Utilidad columns → price_usd = "Precio de Venta"
2. Servicios/Flete (ARS): no Costo column → price_ars = "Precio de Venta"
3. CSV: any format with headers that can be mapped to standard fields

The parser:
- Detects format automatically from headers
- Extracts price WITHOUT IVA (never uses "Con IVA" column as source)
- Classifies each item into the correct catalog by matching SKU
- Returns structured preview data with diffs against current catalog
"""

import csv
import io
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Format detection ─────────────────────────────────────────────────────────

# Dux XLS header patterns (row 6)
DUX_MATERIALS_HEADERS = {"código", "producto", "costo", "porc. utilidad", "precio de venta"}
DUX_SERVICIOS_HEADERS = {"código", "producto", "precio de venta"}

# Catalog classification: maps SKU → catalog name
# Built dynamically from current catalog data


def detect_format(headers: list[str]) -> str:
    """Detect Dux export format from headers.

    Returns: 'dux_materials_usd' | 'dux_servicios_ars' | 'csv_generic'
    """
    normalized = {h.lower().strip() for h in headers if h}
    if "costo" in normalized and "porc. utilidad" in normalized:
        return "dux_materials_usd"
    if "código" in normalized and "precio de venta" in normalized:
        return "dux_servicios_ars"
    return "csv_generic"


# ── File reading ─────────────────────────────────────────────────────────────

def read_file(file_bytes: bytes, filename: str) -> tuple[list[str], list[list]]:
    """Read file and return (headers, rows).

    Supports .xls (xlrd), .xlsx (openpyxl), .csv.
    For Dux .xls files, skips metadata rows (0-5) and reads from row 6 (headers).
    """
    ext = Path(filename).suffix.lower()

    if ext == ".xls":
        return _read_xls(file_bytes)
    elif ext == ".xlsx":
        return _read_xlsx(file_bytes)
    elif ext == ".csv":
        return _read_csv(file_bytes)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Usar .xls, .xlsx o .csv")


def _read_xls(data: bytes) -> tuple[list[str], list[list]]:
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)

    # Dux XLS: row 6 = headers, row 7+ = data
    # Detect header row by looking for "Código" in first cell
    header_row = None
    for r in range(min(10, ws.nrows)):
        val = str(ws.cell_value(r, 0)).strip().lower()
        if val in ("código", "codigo", "sku", "code"):
            header_row = r
            break
    if header_row is None:
        header_row = 0  # fallback

    headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(header_row + 1, ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            val = cell.value
            # Convert xlrd dates to string
            if cell.ctype == 3 and val:
                try:
                    val = xlrd.xldate_as_datetime(val, wb.datemode).strftime("%d/%m/%Y")
                except Exception:
                    pass
            row.append(val)
        # Skip completely empty rows
        if any(str(v).strip() for v in row):
            rows.append(row)
    return headers, rows


def _read_xlsx(data: bytes) -> tuple[list[str], list[list]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    # Same header detection as XLS
    all_rows = list(ws.iter_rows(values_only=True))
    header_row_idx = 0
    for i, row in enumerate(all_rows[:10]):
        val = str(row[0] or "").strip().lower()
        if val in ("código", "codigo", "sku", "code"):
            header_row_idx = i
            break

    headers = [str(v or "").strip() for v in all_rows[header_row_idx]]
    rows = []
    for row in all_rows[header_row_idx + 1:]:
        vals = list(row)
        if any(str(v or "").strip() for v in vals):
            rows.append(vals)
    wb.close()
    return headers, rows


def _read_csv(data: bytes) -> tuple[list[str], list[list]]:
    text = data.decode("utf-8-sig")  # Handle BOM
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return [], []
    return all_rows[0], all_rows[1:]


# ── Price extraction ─────────────────────────────────────────────────────────

def _find_col(headers: list[str], *names: str) -> Optional[int]:
    """Find column index by normalized name matching."""
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        for name in names:
            if name in hl:
                return i
    return None


def _parse_number(val) -> Optional[float]:
    """Parse a number from various formats."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else 0.0
    s = str(val).strip().replace("$", "").replace(" ", "")
    # Handle comma as decimal separator (Argentine format)
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "," in s and "." in s:
        s = s.replace(",", "")  # Thousands separator
    try:
        return float(s)
    except ValueError:
        return None


# ── Item extraction ──────────────────────────────────────────────────────────

def extract_items(headers: list[str], rows: list[list], fmt: str) -> list[dict]:
    """Extract normalized items from parsed rows.

    Returns list of: {sku, name, price_no_vat, price_with_vat, currency, last_updated}
    price_no_vat is the authoritative price (Precio de Venta, NOT Con IVA).
    """
    col_sku = _find_col(headers, "código", "codigo", "sku", "code")
    col_name = _find_col(headers, "producto", "nombre", "name", "descripcion")
    col_date = _find_col(headers, "ultima modificacion", "last_updated", "fecha")

    if col_sku is None:
        raise ValueError("No se encontró columna de SKU/Código en el archivo")

    if fmt == "dux_materials_usd":
        # Precio de Venta (col 4) = USD sin IVA — source of truth
        col_price = _find_col(headers, "precio de venta")
        col_price_iva = _find_col(headers, "precio de venta con iva")
        currency = "USD"
    elif fmt == "dux_servicios_ars":
        # Precio de Venta (col 2) = ARS sin IVA — source of truth
        col_price = _find_col(headers, "precio de venta")
        col_price_iva = _find_col(headers, "precio de venta con iva")
        currency = "ARS"
    else:
        # CSV generic: try standard field names
        col_price = _find_col(headers, "precio sin iva", "price_no_vat", "precio_sin_iva",
                              "price_usd", "price_ars", "precio")
        col_price_iva = _find_col(headers, "precio con iva", "price_with_vat", "precio_con_iva")
        currency = "USD"  # Will be overridden by catalog match

    if col_price is None:
        raise ValueError("No se encontró columna de precio sin IVA. "
                         "Columnas disponibles: " + ", ".join(headers))

    items = []
    for row in rows:
        if len(row) <= col_sku:
            continue
        sku = str(row[col_sku] or "").strip()
        if not sku:
            continue

        name = str(row[col_name] or "").strip() if col_name is not None and len(row) > col_name else ""
        price_no_vat = _parse_number(row[col_price]) if len(row) > col_price else None
        price_with_vat = _parse_number(row[col_price_iva]) if col_price_iva is not None and len(row) > col_price_iva else None
        last_updated = str(row[col_date] or "").strip() if col_date is not None and len(row) > col_date else None

        items.append({
            "sku": sku,
            "name": name,
            "price_no_vat": price_no_vat,
            "price_with_vat": price_with_vat,
            "currency": currency,
            "last_updated": last_updated,
        })

    return items


# ── Catalog classification ───────────────────────────────────────────────────

def build_sku_index(catalogs: dict[str, list]) -> dict[str, str]:
    """Build {sku → catalog_name} index from all current catalogs."""
    index = {}
    for cat_name, items in catalogs.items():
        if not isinstance(items, list):
            continue
        for item in items:
            if isinstance(item, dict) and "sku" in item:
                index[item["sku"].upper()] = cat_name
    return index


def classify_items(items: list[dict], sku_index: dict[str, str]) -> dict[str, list[dict]]:
    """Classify items by target catalog using SKU matching.

    Returns: {catalog_name: [items], "_unmatched": [items]}
    """
    classified: dict[str, list[dict]] = {"_unmatched": []}
    for item in items:
        sku_upper = item["sku"].upper()
        cat = sku_index.get(sku_upper)
        if cat:
            classified.setdefault(cat, [])
            classified[cat].append(item)
        else:
            classified["_unmatched"].append(item)
    return classified


# ── Diff generation ──────────────────────────────────────────────────────────

def generate_diff(catalog_name: str, current_items: list[dict], new_items: list[dict]) -> dict:
    """Generate diff between current catalog and incoming items.

    Returns: {
        catalog: str,
        updated: [{sku, name, old_price, new_price, change_pct}],
        new: [{sku, name, price}],
        missing: [{sku, name, price}],  # in catalog but not in import
        zero_price: [{sku, name}],       # items with $0 price
        warnings: [str],
    }
    """
    # Determine price field
    is_usd = any(item.get("currency") == "USD" for item in new_items[:5])
    price_field = "price_usd" if is_usd else "price_ars"

    current_by_sku = {}
    for item in current_items:
        if isinstance(item, dict) and "sku" in item:
            current_by_sku[item["sku"].upper()] = item

    updated = []
    new = []
    zero_price = []
    unchanged = 0
    import_skus = set()
    warnings = []

    for item in new_items:
        sku = item["sku"]
        sku_upper = sku.upper()
        import_skus.add(sku_upper)
        price = item.get("price_no_vat")

        if price is not None and price == 0:
            zero_price.append({"sku": sku, "name": item.get("name", "")})
            continue

        if sku_upper in current_by_sku:
            cur = current_by_sku[sku_upper]
            old_price = cur.get(price_field, 0)
            if price is not None and old_price and abs(price - old_price) > 0.01:
                change_pct = round((price - old_price) / old_price * 100, 1) if old_price else 0
                entry = {
                    "sku": sku,
                    "name": item.get("name") or cur.get("name", ""),
                    "old_price": old_price,
                    "new_price": price,
                    "change_pct": change_pct,
                }
                if abs(change_pct) > 30:
                    warnings.append(f"SKU '{sku}': cambio de precio {change_pct:+.1f}% ({old_price} → {price})")
                updated.append(entry)
            else:
                unchanged += 1
        else:
            if price is not None and price > 0:
                new.append({"sku": sku, "name": item.get("name", ""), "price": price})

    # Missing: in current but not in import
    missing = []
    for sku_upper, item in current_by_sku.items():
        if sku_upper not in import_skus:
            missing.append({
                "sku": item.get("sku", ""),
                "name": item.get("name", ""),
                "price": item.get(price_field, 0),
            })

    if len(missing) > len(current_by_sku) * 0.5 and len(current_by_sku) > 3:
        warnings.append(
            f"El archivo no incluye {len(missing)} de {len(current_by_sku)} items actuales. "
            f"Esto es normal si el archivo es parcial. Los faltantes NO se borran."
        )

    return {
        "catalog": catalog_name,
        "currency": "USD" if is_usd else "ARS",
        "price_field": price_field,
        "updated": updated,
        "new": new,
        "missing": missing,
        "zero_price": zero_price,
        "unchanged": unchanged,
        "warnings": warnings,
        "total_in_file": len(new_items),
        "total_in_catalog": len(current_by_sku),
    }


# ── Full pipeline ────────────────────────────────────────────────────────────

def parse_import_file(file_bytes: bytes, filename: str, current_catalogs: dict[str, list]) -> dict:
    """Full pipeline: read → detect → extract → classify → diff.

    Args:
        file_bytes: raw file content
        filename: original filename for extension detection
        current_catalogs: {catalog_name: [items]} for all catalogs

    Returns: {
        format: str,
        total_items: int,
        catalogs: {
            catalog_name: {diff data}
        },
        unmatched: [{sku, name, price}],
        iva_warning: bool,  # True if only "con IVA" column found
        warnings: [str],
    }
    """
    # 1. Read file
    headers, rows = read_file(file_bytes, filename)

    # 2. Detect format
    fmt = detect_format(headers)
    logging.info(f"[import] Detected format: {fmt} from {filename} ({len(rows)} rows)")

    # 3. Check IVA safety
    has_price_col = _find_col(headers, "precio de venta") is not None
    has_iva_col = _find_col(headers, "precio de venta con iva", "precio con iva") is not None
    iva_warning = False
    if has_iva_col and not has_price_col:
        iva_warning = True
        logging.warning(f"[import] Only 'con IVA' column found — requires explicit confirmation")

    # 4. Extract items
    items = extract_items(headers, rows, fmt)
    logging.info(f"[import] Extracted {len(items)} items")

    # 5. Build SKU index and classify
    sku_index = build_sku_index(current_catalogs)
    classified = classify_items(items, sku_index)

    # 6. Generate diffs per catalog
    catalog_diffs = {}
    global_warnings = []

    for cat_name, cat_items in classified.items():
        if cat_name == "_unmatched":
            continue
        current = current_catalogs.get(cat_name, [])
        if isinstance(current, dict):
            current = current.get("items", [])
        diff = generate_diff(cat_name, current, cat_items)
        catalog_diffs[cat_name] = diff
        global_warnings.extend(diff["warnings"])

    unmatched = [
        {"sku": i["sku"], "name": i.get("name", ""), "price": i.get("price_no_vat")}
        for i in classified.get("_unmatched", [])
    ]
    if unmatched:
        global_warnings.append(
            f"{len(unmatched)} items no matchean con ningún catálogo existente. "
            f"Requieren asignación manual si son válidos."
        )

    return {
        "format": fmt,
        "total_items": len(items),
        "catalogs": catalog_diffs,
        "unmatched": unmatched,
        "iva_warning": iva_warning,
        "warnings": global_warnings,
    }
