import math
import shutil
import logging
from pathlib import Path
from datetime import datetime
import asyncio

BASE_DIR = Path(__file__).parent.parent.parent.parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"

from app.core.static import OUTPUT_DIR


DELIVERY_SUFFIX = "días desde la toma de medidas"
from app.core.catalog_dir import CATALOG_DIR

_company_config_cache: dict | None = None


def _load_company_config() -> dict:
    """Load company + conditions from config.json. Cached in memory."""
    global _company_config_cache
    if _company_config_cache is not None:
        return _company_config_cache

    import json
    try:
        with open(CATALOG_DIR / "config.json") as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}

    company = cfg.get("company", {})
    conditions = cfg.get("conditions", {})

    _company_config_cache = {
        "name": company.get("name", "D'ANGELO"),
        "subtitle": company.get("subtitle", "MARMOLERIA"),
        "address": company.get("address", "SAN NICOLAS 1160"),
        "phone": company.get("phone", "341-3082996"),
        "email": company.get("email", "marmoleriadangelo@gmail.com"),
        "conditions_general": conditions.get("general", ""),
        "conditions_payment": conditions.get("payment", ""),
    }
    return _company_config_cache


def invalidate_company_config_cache():
    """Clear cached company config so next call reloads from disk."""
    global _company_config_cache
    _company_config_cache = None


# ── Unicode → latin-1 sanitizer for FPDF helvetica ──────────────────────────
# FPDF's default helvetica font is latin-1 only. Em/en dashes, smart quotes,
# ellipsis, etc. crash with UnicodeEncodeError. Normalize everything we push
# into PDF cells so the generator never blows up on punctuation variants that
# the LLM or operator copy-paste into names/projects.
_PDF_CHAR_MAP = {
    "\u2014": "-",   # em dash —
    "\u2013": "-",   # en dash –
    "\u2212": "-",   # minus sign −
    "\u2018": "'",   # left single quote ‘
    "\u2019": "'",   # right single quote ’
    "\u201C": '"',   # left double quote “
    "\u201D": '"',   # right double quote ”
    "\u2026": "...", # ellipsis …
    "\u00A0": " ",   # non-breaking space
    "\u2022": "-",   # bullet •
    "\u00B7": "-",   # middle dot ·
    "\u2009": " ",   # thin space
    "\u200B": "",    # zero-width space
    "\u2190": "<-",
    "\u2192": "->",
}


def _pdf_safe(text) -> str:
    """Return a latin-1-safe version of text for FPDF helvetica. Non-str inputs
    are coerced to str; unknown non-latin-1 chars are replaced with '?'."""
    if text is None:
        return ""
    s = str(text)
    for src, dst in _PDF_CHAR_MAP.items():
        if src in s:
            s = s.replace(src, dst)
    # Final safety net: drop anything else outside latin-1
    try:
        s.encode("latin-1")
        return s
    except UnicodeEncodeError:
        return s.encode("latin-1", errors="replace").decode("latin-1")


def _make_safe_fpdf():
    """Return an FPDF subclass whose text-writing methods auto-sanitize Unicode.

    This is the root-cause fix for 'UnicodeEncodeError: ... in helvetica' crashes
    triggered by em dashes / smart quotes / etc. that operators or the LLM inject
    into client names, project titles, piece descriptions, etc.
    """
    from fpdf import FPDF as _BaseFPDF

    class _SafeFPDF(_BaseFPDF):
        def cell(self, *args, **kwargs):
            if "txt" in kwargs:
                kwargs["txt"] = _pdf_safe(kwargs["txt"])
            elif "text" in kwargs:
                kwargs["text"] = _pdf_safe(kwargs["text"])
            if len(args) >= 3 and isinstance(args[2], str):
                args = list(args)
                args[2] = _pdf_safe(args[2])
                args = tuple(args)
            return super().cell(*args, **kwargs)

        def multi_cell(self, *args, **kwargs):
            if "txt" in kwargs:
                kwargs["txt"] = _pdf_safe(kwargs["txt"])
            elif "text" in kwargs:
                kwargs["text"] = _pdf_safe(kwargs["text"])
            if len(args) >= 3 and isinstance(args[2], str):
                args = list(args)
                args[2] = _pdf_safe(args[2])
                args = tuple(args)
            return super().multi_cell(*args, **kwargs)

        def text(self, x, y, txt):
            return super().text(x, y, _pdf_safe(txt))

    return _SafeFPDF


# ── Planilla de Cómputo (m² override) footnote ──────────────────────────────
# Rendered in both PDF and Excel whenever any piece in the quote declared an
# m2_override (operator pulled the value from the comitente's Planilla de
# Cómputo and D'Angelo cotizes without recomputing from largo×prof).
_M2_OVERRIDE_FOOTNOTE = (
    "Los m² de las piezas marcadas con (*) corresponden a valores declarados "
    "en la Planilla de Cómputo del comitente, los cuales incluyen superficies "
    "de zócalo y/o frente según especificación técnica del proyecto. "
    "D'Angelo Marmolería cotiza en base a dichos valores sin modificación."
)


def _pdf_has_m2_override(data: dict) -> bool:
    """Return True if the renderer should show the Planilla-de-Cómputo footnote.

    Two signal paths — we accept either:
      1. Explicit flag `has_m2_override` on the top-level data dict (set by
         the calculator).
      2. Fallback: any piece label in `sectors[].pieces` ends with " *".
    """
    if data.get("has_m2_override"):
        return True
    for sector in data.get("sectors") or []:
        for piece in sector.get("pieces") or []:
            if isinstance(piece, str) and piece.rstrip().endswith("*"):
                return True
    return False


def _fmt_ars(value: float) -> str:
    """Format ARS price: $65.147,34 (dot thousands, comma decimal, 2 decimals)."""
    n = round(value, 2)
    # Format with 2 decimals, then swap separators for Argentine locale
    raw = f"{abs(n):,.2f}"  # "65,147.34"
    # Swap: comma→temp, dot→comma, temp→dot
    formatted = raw.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"${formatted}" if n >= 0 else f"-${formatted}"


def _fmt_usd(value: float) -> str:
    """Format USD price: USD 1.937 (dot for thousands, no decimals)."""
    n = round(value)
    formatted = f"{abs(n):,}".replace(",", ".")
    return f"USD {formatted}" if n >= 0 else f"-USD {formatted}"


def _fmt_qty(value: float) -> str:
    """Format quantity Argentine style: 1,20 (comma for decimal)."""
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}".replace(".", ",")


def _normalize_delivery(raw: str) -> str:
    """Ensure delivery text is complete, not just a number or short phrase.

    Cases handled:
    - "20"          → "20 días desde la toma de medidas"
    - "20 dias"     → "20 días desde la toma de medidas"   (PR #33 fix)
    - "20 días"     → "20 días desde la toma de medidas"
    - "A confirmar" → left as-is
    - "30 dias desde la toma de medidas en obra" → left as-is (already full)
    """
    if not raw:
        return ""
    raw = raw.strip()
    # If it's just a number like "30", add the full text
    if raw.isdigit():
        return f"{raw} {DELIVERY_SUFFIX}"
    # If it's "N dias" or "N días" (short form), complete it.
    import re as _re_deliv
    _short = _re_deliv.match(r'^\s*(\d{1,3})\s*d[ií]as\s*$', raw, _re_deliv.IGNORECASE)
    if _short:
        return f"{_short.group(1)} {DELIVERY_SUFFIX}"
    # Already contains "toma de medidas" or similar → leave as is.
    return raw


def _strip_duplicate_dims_in_labels(quote_data: dict) -> None:
    """PR #48 — limpia dimensiones duplicadas en piece labels al vuelo.

    El builder de labels en calculator.py antepone "<largo> × <prof>" al
    description de cada pieza. Si Valentina (o un breakdown legacy guardado
    pre-fix) ya metió las dimensiones dentro del description (ej:
    "ME01-B Mesada recta - 2.15m × 0.60m c/zócalo h:10cm"), quedan dobles
    en el PDF.

    Este strip corre en el renderer para que /regenerate también limpie
    presupuestos legacy sin re-ejecutar calculate_quote.

    Muta `quote_data["sectors"][*]["pieces"]` en su lugar.
    """
    import re as _re_strip
    _pat = _re_strip.compile(
        r'\s*[-–—]\s*\d+[.,]?\d*\s*m\s*[×xX]\s*\d+[.,]?\d*\s*m\s*',
        _re_strip.IGNORECASE,
    )
    for sector in quote_data.get("sectors") or []:
        pieces = sector.get("pieces") or []
        cleaned = []
        for p in pieces:
            if isinstance(p, str):
                _c = _pat.sub(' ', p)
                _c = _re_strip.sub(r'\s{2,}', ' ', _c).strip()
                cleaned.append(_c)
            else:
                cleaned.append(p)
        sector["pieces"] = cleaned


async def generate_documents(quote_id: str, quote_data: dict) -> dict:
    """Generate PDF and Excel for a quote."""
    try:
        # PR #48 — strip defensivo de dimensiones duplicadas en piece labels.
        # Aplica acá (entry point de render) para que /regenerate también
        # limpie breakdowns legacy que tienen labels duplicados guardados.
        _strip_duplicate_dims_in_labels(quote_data)

        # PR #40 — reemplazar delivery_days vago (A confirmar, A convenir,
        # N/A, vacío, -) con el default del config.json. Se aplica acá en
        # vez de solo en _validate_quote_data para que el flujo /regenerate
        # (que NO pasa por el validator) también lo corrija.
        _current = (quote_data.get("delivery_days") or "").strip()
        _VAGUE = {"", "a confirmar", "a convenir", "n/a", "n/d", "-", "—", "."}
        if _current.lower() in _VAGUE:
            try:
                import json as _json
                from pathlib import Path as _P
                _cfg_path = _P(__file__).resolve().parent.parent.parent.parent / "catalog" / "config.json"
                _cfg = _json.loads(_cfg_path.read_text(encoding="utf-8"))
                quote_data["delivery_days"] = _cfg.get("delivery_days", {}).get(
                    "display", "30 dias desde la toma de medidas"
                )
            except Exception:
                quote_data["delivery_days"] = "30 dias desde la toma de medidas"

        client_name = quote_data["client_name"]
        material = quote_data.get("material_name", "")
        prefix = quote_data.get("filename_prefix", "")
        # Always use server date — never trust Claude's date
        date_str = datetime.now().strftime("%d.%m.%Y")
        # Sanitize filename: remove any remaining invalid characters
        filename_base = f"{prefix}{client_name} - {material} - {date_str}"
        filename_base = filename_base.replace("/", "-").replace("\\", "-")

        quote_dir = OUTPUT_DIR / quote_id
        quote_dir.mkdir(exist_ok=True)

        pdf_path = quote_dir / f"{filename_base}.pdf"
        excel_path = quote_dir / f"{filename_base}.xlsx"

        # Generate both in parallel — run blocking I/O in thread pool
        await asyncio.gather(
            asyncio.to_thread(_generate_excel, excel_path, quote_data),
            asyncio.to_thread(_generate_pdf, pdf_path, quote_data),
        )

        return {
            "ok": True,
            "pdf_url": f"/files/{quote_id}/{filename_base}.pdf",
            "excel_url": f"/files/{quote_id}/{filename_base}.xlsx",
            "filename_base": filename_base,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


async def generate_edificio_documents(
    quote_id: str,
    paso2_calc: dict,
    summary: dict,
    client_name: str = "",
    project: str = "",
    localidad: str = "Rosario",
) -> dict:
    """Generate 3 PDF+Excel for edificio using build_edificio_doc_context.

    Presentation model built by edificio_parser, not ad-hoc here.
    Uses _generate_pdf/_generate_excel (original D'Angelo template).
    """
    from app.modules.quote_engine.edificio_parser import build_edificio_doc_context

    date_str = datetime.now().strftime("%d.%m.%Y")
    quote_dir = OUTPUT_DIR / quote_id
    quote_dir.mkdir(exist_ok=True)

    contexts = build_edificio_doc_context(summary, paso2_calc, client_name, project)
    generated = []

    for ctx in contexts:
        mat_name = ctx.get("_mat_name_raw", "")
        cur = ctx.get("_currency", "ARS")

        safe_mat = (ctx.get("material_name") or mat_name).replace("/", "-").replace("\\", "-")
        base_name = f"{client_name} - {safe_mat} - {date_str}"
        pdf_path = quote_dir / f"{base_name}.pdf"
        excel_path = quote_dir / f"{base_name}.xlsx"

        # Use edificio-specific renderers that respect show_mo and grand_total_text
        await asyncio.gather(
            asyncio.to_thread(_generate_edificio_pdf, pdf_path, ctx),
            asyncio.to_thread(_generate_edificio_excel, excel_path, ctx),
        )

        generated.append({
            "material": mat_name or ctx.get("material_name", ""),
            "pdf_url": f"/files/{quote_id}/{base_name}.pdf",
            "excel_url": f"/files/{quote_id}/{base_name}.xlsx",
            "total_ars": ctx.get("total_ars", 0),
            "total_usd": ctx.get("total_usd", 0),
            "currency": cur,
        })

    # ── 4th document: Resumen General de Obra ──
    resumen_name = f"{client_name} - Resumen General de Obra - {date_str}"
    resumen_pdf = quote_dir / f"{resumen_name}.pdf"
    resumen_xlsx = quote_dir / f"{resumen_name}.xlsx"

    # Build material rows from paso2_calc (read-only, no recomputation)
    calc_results = paso2_calc.get("calc_results", {})
    mat_rows = []
    total_mat_ars = 0
    total_mat_usd = 0
    mat_display_names = []
    for mat_name_r, mr in calc_results.items():
        if not mr.get("ok"):
            continue
        cur = mr["currency"]
        net = mr["material_net"]
        mat_rows.append({
            "name": mr.get("catalog_name", mat_name_r),
            "m2": mr["m2"],
            "currency": cur,
            "price_unit": mr["price_unit"],
            "subtotal": mr["material_total"],
            "discount_pct": mr.get("discount_pct", 0),
            "discount_amount": mr.get("discount_amount", 0),
            "net": net,
        })
        mat_display_names.append(mr.get("catalog_name", mat_name_r))
        if cur == "ARS":
            total_mat_ars += net
        else:
            total_mat_usd += net

    resumen_data = {
        "client_name": client_name,
        "project": project,
        "materials": mat_rows,
        "mo_items": paso2_calc.get("mo_items", []),
        "mo_total": paso2_calc.get("mo_total", 0),
        "total_mat_ars": total_mat_ars,
        "total_mat_usd": total_mat_usd,
        "grand_total_ars": paso2_calc.get("grand_total_ars", 0),
        "grand_total_usd": paso2_calc.get("grand_total_usd", 0),
        "material_names": mat_display_names,
    }

    await asyncio.gather(
        asyncio.to_thread(_generate_resumen_obra_pdf, resumen_pdf, resumen_data),
        asyncio.to_thread(_generate_resumen_obra_excel, resumen_xlsx, resumen_data),
    )

    generated.append({
        "material": "Resumen General de Obra",
        "pdf_url": f"/files/{quote_id}/{resumen_name}.pdf",
        "excel_url": f"/files/{quote_id}/{resumen_name}.xlsx",
        "total_ars": paso2_calc.get("grand_total_ars", 0),
        "total_usd": paso2_calc.get("grand_total_usd", 0),
        "currency": "consolidado",
    })

    return {
        "ok": True,
        "generated": generated,
        "grand_total_ars": paso2_calc.get("grand_total_ars", 0),
        "grand_total_usd": paso2_calc.get("grand_total_usd", 0),
    }


def _generate_resumen_obra_pdf(pdf_path: Path, data: dict) -> None:
    """Generate consolidated project summary PDF — same D'Angelo look."""
    FPDF = _make_safe_fpdf()
    from app.modules.quote_engine.edificio_parser import _fmt_num

    co = _load_company_config()
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Top bar
    pdf.set_fill_color(26, 47, 94)
    pdf.rect(0, 0, 210, 4, "F")

    # Logo
    logo_path = TEMPLATES_DIR / "logo-dangelo.png"
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=10, w=55)
        pdf.set_y(35)
    else:
        pdf.set_y(12)
        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 12, co["name"], new_x="LMARGIN", new_y="NEXT")

    # Header
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"{co['address']} | Tel: {co['phone']} | {co['email']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    col_w = 95
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(col_w, 5, f"Cliente: {data['client_name']}")
    pdf.cell(col_w, 5, "Forma de pago", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(col_w, 5, "")
    pdf.cell(col_w, 5, "CONTADO EFVO", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(col_w, 5, f"Proyecto: {data['project']}")
    pdf.cell(col_w, 5, "Fecha de entrega", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(col_w, 5, "")
    _plazo = data.get("plazo") or "Segun cronograma de obra"
    pdf.cell(col_w, 5, _plazo, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(col_w, 5, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
    pdf.ln(5)

    pdf.set_draw_color(26, 26, 26)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    # Title
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "RESUMEN GENERAL DE OBRA", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ── 1. Materials table ──
    pdf.set_font("Helvetica", "B", 8)
    mw = [52, 12, 12, 22, 24, 22, 24]  # widths
    headers = ["Material", "m2", "Mon", "Precio unit", "Subtotal", "Descuento", "Total neto"]
    for i, h in enumerate(headers):
        pdf.cell(mw[i], 5, h, align="R" if i > 0 else "L")
    pdf.ln()
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(1)

    pdf.set_font("Helvetica", "", 8)
    for m in data["materials"]:
        cur = m["currency"]
        sym = "USD " if cur == "USD" else "$"
        pdf.cell(mw[0], 5, m["name"][:30])
        pdf.cell(mw[1], 5, _fmt_num(m["m2"]), align="R")
        pdf.cell(mw[2], 5, cur, align="R")
        pdf.cell(mw[3], 5, f"{sym}{_fmt_num(m['price_unit'], 0)}", align="R")
        pdf.cell(mw[4], 5, f"{sym}{_fmt_num(m['subtotal'], 0)}", align="R")
        if m["discount_pct"]:
            pdf.cell(mw[5], 5, f"-{sym}{_fmt_num(m['discount_amount'], 0)}", align="R")
        else:
            pdf.cell(mw[5], 5, "-", align="R")
        pdf.cell(mw[6], 5, f"{sym}{_fmt_num(m['net'], 0)}", align="R")
        pdf.ln()

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 9)
    if data["total_mat_ars"]:
        pdf.cell(0, 5, f"Total materiales ARS: ${_fmt_num(data['total_mat_ars'], 0)}", new_x="LMARGIN", new_y="NEXT")
    if data["total_mat_usd"]:
        pdf.cell(0, 5, f"Total materiales USD: USD {_fmt_num(data['total_mat_usd'], 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 2. MO table ──
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, "MANO DE OBRA GLOBAL", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    ow = [80, 25, 35, 35]
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(ow[0], 5, "Concepto")
    pdf.cell(ow[1], 5, "Cantidad", align="R")
    pdf.cell(ow[2], 5, "Precio c/IVA", align="R")
    pdf.cell(ow[3], 5, "Total", align="R")
    pdf.ln()
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(1)

    pdf.set_font("Helvetica", "", 8)
    for mo in data["mo_items"]:
        qty = mo.get("qty", 1)
        qty_str = _fmt_num(qty, 2) if isinstance(qty, float) else str(qty)
        pdf.cell(ow[0], 5, mo.get("desc", ""))
        pdf.cell(ow[1], 5, qty_str, align="R")
        pdf.cell(ow[2], 5, f"${_fmt_num(mo.get('price', 0), 0)}", align="R")
        pdf.cell(ow[3], 5, f"${_fmt_num(mo.get('total', 0), 0)}", align="R")
        pdf.ln()

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, f"Total mano de obra ARS: ${_fmt_num(data['mo_total'], 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 3. Grand totals ──
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "TOTALES GENERALES DEL PROYECTO", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)
    pdf.set_font("Helvetica", "", 9)
    if data["total_mat_ars"]:
        pdf.cell(0, 5, f"Total materiales ARS: ${_fmt_num(data['total_mat_ars'], 0)}", new_x="LMARGIN", new_y="NEXT")
    if data["total_mat_usd"]:
        pdf.cell(0, 5, f"Total materiales USD: USD {_fmt_num(data['total_mat_usd'], 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Total mano de obra ARS: ${_fmt_num(data['mo_total'], 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, f"TOTAL FINAL ARS: ${_fmt_num(data['grand_total_ars'], 0)}", new_x="LMARGIN", new_y="NEXT")
    if data["grand_total_usd"]:
        pdf.cell(0, 6, f"TOTAL FINAL USD: USD {_fmt_num(data['grand_total_usd'], 0)}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 4. Clarification ──
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 4, "Este documento resume el total consolidado de la obra.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, "Ademas, se emiten presupuestos individuales por material:", new_x="LMARGIN", new_y="NEXT")
    for name in data.get("material_names", []):
        pdf.cell(0, 4, f"  - {name}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ── 4b. Operator notes (optional) ──
    _notes = (data.get("notes") or "").strip()
    if _notes:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, "NOTAS ADICIONALES", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(180, 4, _notes, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # ── 5. Conditions ──
    pdf.set_font("Helvetica", "", 7)
    pdf.multi_cell(180, 3.5, "Forma de pago: Contado", new_x="LMARGIN", new_y="NEXT")
    if co.get("conditions_general"):
        pdf.multi_cell(180, 3.5, co["conditions_general"].strip(), new_x="LMARGIN", new_y="NEXT")
    if co.get("conditions_payment"):
        pdf.ln(1)
        pdf.multi_cell(180, 3.5, co["conditions_payment"].strip(), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 7)
    pdf.multi_cell(180, 4, "No se suben mesadas que no entren en ascensor", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(pdf_path))


def _generate_resumen_obra_excel(excel_path: Path, data: dict) -> None:
    """Generate consolidated project summary Excel — one sheet, mirrors PDF."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from app.modules.quote_engine.edificio_parser import _fmt_num

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Obra"

    bold = Font(name="Arial", bold=True, size=10)
    bold_big = Font(name="Arial", bold=True, size=12)
    normal = Font(name="Arial", size=10)
    small = Font(name="Arial", size=9)
    italic = Font(name="Arial", italic=True, size=9)
    ars_fmt = '"$"#,##0'

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    r = 1
    ws.cell(r, 1, "D'ANGELO MARMOLERIA").font = Font(name="Arial", bold=True, size=14)
    r = 3
    ws.cell(r, 1, f"Cliente: {data['client_name']}").font = bold
    ws.cell(r, 4, "Forma de pago: CONTADO EFVO").font = normal
    r = 4
    ws.cell(r, 1, f"Proyecto: {data['project']}").font = bold
    ws.cell(r, 4, "Fecha de entrega: Segun cronograma de obra").font = normal
    r = 5
    ws.cell(r, 1, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}").font = normal
    r = 7
    ws.cell(r, 1, "RESUMEN GENERAL DE OBRA").font = bold_big

    # Materials table
    r = 9
    for ci, h in enumerate(["Material", "m2", "Moneda", "Precio unit", "Subtotal", "Descuento", "Total neto"], 1):
        ws.cell(r, ci, h).font = bold
    r = 10
    for m in data["materials"]:
        cur = m["currency"]
        ws.cell(r, 1, m["name"]).font = normal
        ws.cell(r, 2, m["m2"]).font = normal
        ws.cell(r, 2).number_format = 'General'
        ws.cell(r, 3, cur).font = normal
        if cur == "USD":
            ws.cell(r, 4, _fmt_usd(m['price_unit'])).font = normal
            ws.cell(r, 5, _fmt_usd(m['subtotal'])).font = normal
            ws.cell(r, 6, f"-{_fmt_usd(m['discount_amount'])}" if m['discount_amount'] else "-").font = normal
            ws.cell(r, 7, _fmt_usd(m['net'])).font = bold
        else:
            ws.cell(r, 4, m["price_unit"]).font = normal
            ws.cell(r, 4).number_format = ars_fmt
            ws.cell(r, 5, m["subtotal"]).font = normal
            ws.cell(r, 5).number_format = ars_fmt
            ws.cell(r, 6, -m["discount_amount"] if m["discount_amount"] else 0).font = normal
            ws.cell(r, 6).number_format = ars_fmt
            ws.cell(r, 7, m["net"]).font = bold
            ws.cell(r, 7).number_format = ars_fmt
        r += 1

    r += 1
    if data["total_mat_ars"]:
        ws.cell(r, 1, "Total materiales ARS").font = bold
        ws.cell(r, 7, data["total_mat_ars"]).font = bold
        ws.cell(r, 7).number_format = ars_fmt
        r += 1
    if data["total_mat_usd"]:
        ws.cell(r, 1, "Total materiales USD").font = bold
        ws.cell(r, 7, _fmt_usd(data['total_mat_usd'])).font = bold
        r += 1

    # MO table
    r += 2
    ws.cell(r, 1, "MANO DE OBRA GLOBAL").font = bold_big
    r += 1
    for ci, h in enumerate(["Concepto", "Cantidad", "", "Precio c/IVA", "Total"], 1):
        if h:
            ws.cell(r, ci, h).font = bold
    r += 1
    for mo in data["mo_items"]:
        ws.cell(r, 1, mo.get("desc", "")).font = normal
        ws.cell(r, 2, mo.get("qty", 1)).font = normal
        ws.cell(r, 4, mo.get("price", 0)).font = normal
        ws.cell(r, 4).number_format = ars_fmt
        ws.cell(r, 5, mo.get("total", 0)).font = normal
        ws.cell(r, 5).number_format = ars_fmt
        r += 1

    r += 1
    ws.cell(r, 1, "Total mano de obra ARS").font = bold
    ws.cell(r, 5, data["mo_total"]).font = bold
    ws.cell(r, 5).number_format = ars_fmt

    # Grand totals
    r += 2
    ws.cell(r, 1, "TOTALES GENERALES DEL PROYECTO").font = bold_big
    r += 1
    if data["total_mat_ars"]:
        ws.cell(r, 1, "Total materiales ARS").font = normal
        ws.cell(r, 5, data["total_mat_ars"]).font = normal
        ws.cell(r, 5).number_format = ars_fmt
        r += 1
    if data["total_mat_usd"]:
        ws.cell(r, 1, "Total materiales USD").font = normal
        ws.cell(r, 5, _fmt_usd(data['total_mat_usd'])).font = normal
        r += 1
    ws.cell(r, 1, "Total mano de obra ARS").font = normal
    ws.cell(r, 5, data["mo_total"]).font = normal
    ws.cell(r, 5).number_format = ars_fmt
    r += 1
    ws.cell(r, 1, "TOTAL FINAL ARS").font = bold
    ws.cell(r, 5, data["grand_total_ars"]).font = bold
    ws.cell(r, 5).number_format = ars_fmt
    r += 1
    if data["grand_total_usd"]:
        ws.cell(r, 1, "TOTAL FINAL USD").font = bold
        ws.cell(r, 5, _fmt_usd(data['grand_total_usd'])).font = bold

    # Clarification
    r += 2
    ws.cell(r, 1, "Este documento resume el total consolidado de la obra.").font = italic
    r += 1
    ws.cell(r, 1, "Ademas, se emiten presupuestos individuales por material:").font = italic
    r += 1
    for name in data.get("material_names", []):
        ws.cell(r, 1, f"  - {name}").font = italic
        r += 1

    # Conditions — match PDF
    r += 1
    co = _load_company_config()
    cond_font = Font(name="Arial", size=7)
    ws.cell(r, 1, "Forma de pago: Contado").font = cond_font
    r += 1
    if co.get("conditions_general"):
        for line in co["conditions_general"].split("\n"):
            if line.strip():
                ws.cell(r, 1, line.strip()).font = cond_font
                r += 1
    if co.get("conditions_payment"):
        r += 1
        for line in co["conditions_payment"].split("\n"):
            if line.strip():
                ws.cell(r, 1, line.strip()).font = cond_font
                r += 1
    # Footer
    r += 1
    ws.cell(r, 1, "No se suben mesadas que no entren en ascensor").font = Font(name="Arial", italic=True, size=7)

    wb.save(str(excel_path))
    _force_arial_font(str(excel_path))


def _generate_edificio_pdf(pdf_path: Path, data: dict) -> None:
    """Generate edificio PDF — same D'Angelo look, respects show_mo and grand_total_text."""
    FPDF = _make_safe_fpdf()
    import re as _re

    co = _load_company_config()
    client_name = data.get("client_name", "")
    project = data.get("project", "")
    date_str = datetime.now().strftime("%d/%m/%Y")
    delivery = data.get("delivery_days", "")
    mat_name = data.get("material_name", "")
    mat_m2 = data.get("material_m2", 0)
    mat_price = data.get("material_price_unit", 0)
    currency = data.get("material_currency", "USD")
    discount_pct = data.get("discount_pct", 0)
    sectors = data.get("sectors", [])
    mo_items = data.get("mo_items", [])
    show_mo = data.get("show_mo", len(mo_items) > 0)
    total_ars = data.get("total_ars", 0)
    total_usd = data.get("total_usd", 0)
    grand_text = data.get("grand_total_text", "")
    thickness = data.get("thickness_mm", 20)

    fmt_price = _fmt_usd if currency == "USD" else _fmt_ars

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Top bar
    pdf.set_fill_color(26, 47, 94)
    pdf.rect(0, 0, 210, 4, "F")

    # Logo
    logo_path = TEMPLATES_DIR / "logo-dangelo.png"
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=10, w=55)
        pdf.set_y(35)
    else:
        pdf.set_y(12)
        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 12, co["name"], new_x="LMARGIN", new_y="NEXT")

    # Header — same layout as normal template
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"{co['address']} | Tel: {co['phone']} | {co['email']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    col_w = 95
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(col_w, 5, f"Cliente: {client_name}")
    pdf.cell(col_w, 5, "Forma de pago", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(col_w, 5, "")
    pdf.cell(col_w, 5, "CONTADO EFVO", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(col_w, 5, "Proyecto")
    pdf.cell(col_w, 5, "", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(col_w, 5, project or "")
    pdf.cell(col_w, 5, "", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(col_w, 5, "")
    pdf.cell(col_w, 5, "Fecha de entrega", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(col_w, 5, "")
    pdf.cell(col_w, 5, str(delivery) or "A CONVENIR", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.set_draw_color(26, 26, 26)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    # Table header
    # PR #14 — ampliar columna Descripción (92→112mm) para que entren labels
    # de tipologías edificio largos como "1.45 × 0.50 ME04b-B Mesada recta
    # c/zócalo h:5cm c/frente h:5cm * (×2)". Reduce levemente las columnas
    # de números (~total se mantiene en 190mm).
    w = [112, 18, 30, 30]
    rh = 5
    total_w = sum(w)
    row_n = [0]

    def row_fill():
        fill = row_n[0] % 2 == 1
        if fill:
            pdf.set_fill_color(243, 243, 243)
        return fill

    def row_done():
        row_n[0] += 1

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(w[0], 6, "Descripcion")
    pdf.cell(w[1], 6, "Cantidad", align="R")
    pdf.cell(w[2], 6, "Precio unitario", align="R")
    pdf.cell(w[3], 6, "Precio total", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)

    # Material row
    mat_total_bruto = data.get("material_total") or round(mat_m2 * mat_price)
    _mat_display = f"{mat_name} - {thickness}mm ESPESOR" if not _re.search(r'\d+[Mm][Mm]', mat_name) else mat_name

    f = row_fill()
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(w[0], rh, _mat_display, fill=f)
    pdf.cell(w[1], rh, _fmt_qty(mat_m2), align="R", fill=f)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(w[2], rh, fmt_price(mat_price), align="R", fill=f)
    pdf.cell(w[3], rh, fmt_price(mat_total_bruto), align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
    row_done()

    # Sectors/despiece — first piece row shows DESC and Total
    is_first_piece = True
    for sector in sectors:
        f = row_fill()
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(total_w, rh, sector.get("label", ""), fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()
        raw_pieces = sector.get("pieces", [])
        grouped = []
        seen = {}
        for p in raw_pieces:
            if p in seen:
                seen[p] += 1
            else:
                seen[p] = 1
                grouped.append(p)
        for piece in grouped:
            count = seen[piece]
            display = f"{piece} (x{count})" if count > 1 else piece
            f = row_fill()
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(w[0], rh, display, fill=f)
            if is_first_piece:
                # Show discount + total net on the first piece row
                pdf.set_font("Helvetica", "I", 8)
                if discount_pct:
                    pdf.cell(w[1], rh, "", fill=f)
                    pdf.cell(w[2], rh, f"DESC {discount_pct}%", align="R", fill=f)
                    disc_amount = round(mat_total_bruto * discount_pct / 100)
                    pdf.cell(w[3], rh, fmt_price(disc_amount), align="R", fill=f)
                    pdf.ln()
                    row_done()
                    # Net total row
                    f = row_fill()
                    pdf.cell(w[0], rh, "", fill=f)
                    pdf.set_font("Helvetica", "B", 8)
                    pdf.cell(w[1], rh, "", fill=f)
                    pdf.cell(w[2], rh, f"Total {currency}", align="R", fill=f)
                    mat_net = mat_total_bruto - disc_amount
                    pdf.cell(w[3], rh, fmt_price(mat_net), align="R", fill=f)
                else:
                    pdf.cell(w[1], rh, "", fill=f)
                    pdf.cell(w[2], rh, f"Total {currency}", align="R", fill=f)
                    pdf.cell(w[3], rh, fmt_price(mat_total_bruto), align="R", fill=f)
                is_first_piece = False
            else:
                pdf.cell(w[1] + w[2] + w[3], rh, "", fill=f)
            pdf.ln()
            row_done()

    pdf.ln(3)

    # MO block — only if show_mo
    if show_mo and mo_items:
        f = row_fill()
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(total_w, rh, "MANO DE OBRA", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

        mo_subtotal = 0
        for mo in mo_items:
            f = row_fill()
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(w[0], rh, mo["description"], fill=f)
            pdf.cell(w[1], rh, _fmt_qty(mo["quantity"]), align="R", fill=f)
            pdf.cell(w[2], rh, _fmt_ars(mo["unit_price"]), align="R", fill=f)
            mo_total_line = mo.get("total", round(mo["unit_price"] * mo["quantity"]))
            pdf.cell(w[3], rh, _fmt_ars(mo_total_line), align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
            mo_subtotal += mo_total_line
            row_done()

        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(w[0] + w[1], 5, "")
        pdf.cell(w[2], 5, "TOTAL PESOS", align="R")
        pdf.cell(w[3], 5, _fmt_ars(total_ars), align="R", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(5)

    # Grand total box — use pre-built text from view model
    if grand_text:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_draw_color(26, 26, 26)
        y_box = pdf.get_y()
        pdf.rect(10, y_box, 190, 8)
        pdf.set_xy(10, y_box + 1.5)
        pdf.cell(190, 5, grand_text, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)

    # Conditions — use multi_cell for wrapping long lines
    pdf.set_font("Helvetica", "", 7)
    pdf.multi_cell(180, 3.5, "Forma de pago: Contado", new_x="LMARGIN", new_y="NEXT")
    if co.get("conditions_general"):
        pdf.multi_cell(180, 3.5, co["conditions_general"].strip(), new_x="LMARGIN", new_y="NEXT")
    if co.get("conditions_payment"):
        pdf.ln(1)
        pdf.multi_cell(180, 3.5, co["conditions_payment"].strip(), new_x="LMARGIN", new_y="NEXT")

    # Footer
    pdf.set_font("Helvetica", "I", 7)
    # Planilla de Cómputo footnote — render when any piece used m2_override.
    if _pdf_has_m2_override(data):
        pdf.multi_cell(
            180, 3.5,
            _M2_OVERRIDE_FOOTNOTE,
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(0.5)
    pdf.multi_cell(180, 4, "No se suben mesadas que no entren en ascensor", new_x="LMARGIN", new_y="NEXT")
    # Edificio-only: las piezas se dejan en pie de obra (no se colocan)
    pdf.multi_cell(180, 4, "Las mesadas se dejan en pie de obra", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(pdf_path))


def _generate_edificio_excel(excel_path: Path, data: dict) -> None:
    """Generate edificio Excel — pixel-perfect match with PDF output."""
    import openpyxl
    import re as _re

    TEMPLATE = TEMPLATES_DIR / "excel" / "quote-template-excel.xlsx"
    if not TEMPLATE.exists():
        TEMPLATE = TEMPLATES_DIR / "excel" / "quote-template.xlsx"
    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb.active

    # Unmerge all cells
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    # Remove template conditional formatting (isodd zebra on A23:F36)
    # — we apply our own zebra via PatternFill to match PDF exactly
    ws.conditional_formatting._cf_rules.clear()

    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # ── Font sizes matching PDF exactly ──
    # PDF: table header = bold 9pt, material = bold 9pt, sector = bold 8pt,
    #       piece = normal 8pt, MO header = bold 9pt, MO item = normal 9pt
    bold_9 = Font(name="Arial", bold=True, size=9)
    normal_9 = Font(name="Arial", bold=False, size=9)
    bold_8 = Font(name="Arial", bold=True, size=8)
    normal_8 = Font(name="Arial", bold=False, size=8)
    italic_8 = Font(name="Arial", italic=True, size=8)

    # ARS with 2 decimals matching PDF _fmt_ars → "$65.147,34"
    ars_fmt = '"$"#,##0.00'

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    left_align = Alignment(horizontal="left")

    # Zebra fill — PDF uses rgb(243,243,243) on odd rows
    zebra_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    client_name = data.get("client_name", "")
    project = data.get("project", "")
    date_str = datetime.now().strftime("%d/%m/%Y")
    delivery = data.get("delivery_days", "A CONVENIR")
    mat_name = data.get("material_name", "")
    mat_m2 = data.get("material_m2", 0)
    mat_price = data.get("material_price_unit", 0)
    currency = data.get("material_currency", "USD")
    discount_pct = data.get("discount_pct", 0)
    sectors = data.get("sectors", [])
    mo_items = data.get("mo_items", [])
    show_mo = data.get("show_mo", len(mo_items) > 0)
    total_ars = data.get("total_ars", 0)
    total_usd = data.get("total_usd", 0)
    grand_text = data.get("grand_total_text", "")
    thickness = data.get("thickness_mm", 20)

    fmt_price = _fmt_usd if currency == "USD" else _fmt_ars

    # Header — overwrite template placeholders in their exact cells
    ws["A13"].value = f"Fecha: {date_str}"
    ws["A16"].value = client_name
    ws["C16"].value = "CONTADO EFVO"
    ws["A18"].value = project
    ws["C19"].value = delivery

    # Clear dynamic rows — values, fills, borders, AND hidden state from template
    no_border = Border()
    max_clear = 60 + max(0, len(mo_items)) + len(sectors) * 5
    for row in range(22, max_clear + 1):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border
        ws.row_dimensions[row].height = 15
        ws.row_dimensions[row].hidden = False  # Unhide template row 36

    # Zebra row counter — mirrors PDF row_n[0] counter
    row_n = [0]

    def apply_zebra(sheet_row, cols=6):
        """Apply zebra fill to ALL columns in the row (matching PDF full-width fill)."""
        fill = zebra_fill if row_n[0] % 2 == 1 else no_fill
        for c in range(1, cols + 1):
            ws.cell(sheet_row, c).fill = fill

    def zebra_done():
        row_n[0] += 1

    # Row 22: column headers (PDF: bold 9pt)
    ws.merge_cells("A22:C22")
    ws["A22"].value = "Descripcion"
    ws["A22"].font = bold_9
    ws["A22"].alignment = left_align
    ws["D22"].value = "Cantidad"
    ws["D22"].font = bold_9
    ws["D22"].alignment = right_align
    ws["E22"].value = "Precio unitario"
    ws["E22"].font = bold_9
    ws["E22"].alignment = right_align
    ws["F22"].value = "Precio total"
    ws["F22"].font = bold_9
    ws["F22"].alignment = right_align

    # Row 23: Material (PDF: bold 9pt)
    _mat_display = f"{mat_name} - {thickness}mm ESPESOR" if not _re.search(r'\d+[Mm][Mm]', mat_name) else mat_name
    mat_total_bruto = data.get("material_total") or round(mat_m2 * mat_price)

    apply_zebra(23)
    ws.merge_cells("A23:C23")
    ws["A23"].value = _mat_display
    ws["A23"].font = bold_9
    ws["A23"].alignment = left_align
    ws["D23"].value = _fmt_qty(mat_m2)
    ws["D23"].font = normal_9
    ws["D23"].alignment = right_align
    ws["E23"].value = fmt_price(mat_price)
    ws["E23"].font = normal_9
    ws["E23"].alignment = right_align
    ws["F23"].value = fmt_price(mat_total_bruto)
    ws["F23"].font = normal_9
    ws["F23"].alignment = right_align
    zebra_done()

    # Row 24+: Sectors/despiece + discount/total on FIRST piece row
    r = 24
    first_piece = True
    for sector in sectors:
        apply_zebra(r)
        ws.merge_cells(f"A{r}:C{r}")
        ws.cell(r, 1).value = sector.get("label", "")
        ws.cell(r, 1).font = bold_8
        ws.cell(r, 1).alignment = left_align
        zebra_done()
        r += 1

        raw_pieces = sector.get("pieces", [])
        grouped = []
        seen = {}
        for p in raw_pieces:
            if p in seen:
                seen[p] += 1
            else:
                seen[p] = 1
                grouped.append(p)

        for piece in grouped:
            count = seen[piece]
            display = f"{piece} (x{count})" if count > 1 else piece
            apply_zebra(r)
            ws.merge_cells(f"A{r}:C{r}")
            ws.cell(r, 1).value = display
            ws.cell(r, 1).font = normal_8
            ws.cell(r, 1).alignment = left_align
            if first_piece:
                # Total on same row as first piece (matching PDF)
                if discount_pct:
                    ws.cell(r, 5).value = f"DESC {discount_pct}%"
                    ws.cell(r, 5).font = italic_8
                    ws.cell(r, 5).alignment = right_align
                    disc = round(mat_total_bruto * discount_pct / 100)
                    ws.cell(r, 6).value = fmt_price(disc)
                    ws.cell(r, 6).font = italic_8
                    ws.cell(r, 6).alignment = right_align
                    zebra_done()
                    r += 1
                    apply_zebra(r)
                    ws.cell(r, 5).value = f"Total {currency}"
                    ws.cell(r, 5).font = bold_8
                    ws.cell(r, 5).alignment = right_align
                    mat_net = mat_total_bruto - disc
                    ws.cell(r, 6).value = fmt_price(mat_net)
                    ws.cell(r, 6).font = bold_8
                    ws.cell(r, 6).alignment = right_align
                else:
                    ws.cell(r, 5).value = f"Total {currency}"
                    ws.cell(r, 5).font = bold_8
                    ws.cell(r, 5).alignment = right_align
                    ws.cell(r, 6).value = fmt_price(mat_total_bruto)
                    ws.cell(r, 6).font = bold_8
                    ws.cell(r, 6).alignment = right_align
                first_piece = False
            zebra_done()
            r += 1

    r += 1  # spacer

    # MO block — only if show_mo (PDF: header bold 9pt, items normal 9pt)
    if show_mo and mo_items:
        apply_zebra(r)
        ws.merge_cells(f"A{r}:C{r}")
        ws.cell(r, 1).value = "MANO DE OBRA"
        ws.cell(r, 1).font = bold_9
        ws.cell(r, 1).alignment = left_align
        zebra_done()
        r += 1

        for mo in mo_items:
            apply_zebra(r)
            ws.merge_cells(f"A{r}:C{r}")
            ws.cell(r, 1).value = mo["description"]
            ws.cell(r, 1).font = normal_9
            ws.cell(r, 1).alignment = left_align
            # PR #37 — escribir qty numérica (no string) para que el
            # operador pueda editarla directamente en Google Sheets sin
            # perder decimales. Formato 'General' preserva hasta 4
            # decimales sin trailing zeros.
            ws.cell(r, 4).value = mo["quantity"]
            ws.cell(r, 4).number_format = 'General'
            ws.cell(r, 4).font = normal_9
            ws.cell(r, 4).alignment = right_align
            # MO prices always ARS with 2 decimals
            ws.cell(r, 5).value = _fmt_ars(mo["unit_price"])
            ws.cell(r, 5).font = normal_9
            ws.cell(r, 5).alignment = right_align
            mo_t = mo.get("total", round(mo["unit_price"] * mo["quantity"]))
            ws.cell(r, 6).value = _fmt_ars(mo_t)
            ws.cell(r, 6).font = normal_9
            ws.cell(r, 6).alignment = right_align
            zebra_done()
            r += 1

        r += 1
        ws.cell(r, 5).value = "TOTAL PESOS"
        ws.cell(r, 5).font = bold_9
        ws.cell(r, 5).alignment = right_align
        ws.cell(r, 6).value = _fmt_ars(total_ars)
        ws.cell(r, 6).font = bold_9
        ws.cell(r, 6).alignment = right_align

    r += 2  # spacer

    # Grand total — bordered box, centered (matching PDF rect + align="C")
    if grand_text:
        ws.merge_cells(f"A{r}:F{r}")
        cell_gt = ws.cell(r, 1)
        cell_gt.value = grand_text
        cell_gt.font = Font(name="Arial", bold=True, size=9)
        cell_gt.alignment = center_align
        cell_gt.border = thin_border
        for col in range(2, 7):
            ws.cell(r, col).border = thin_border

    # Conditions — match PDF 7pt, merge A:F + wrap so text flows full width
    r += 2
    co = _load_company_config()
    cond_font = Font(name="Arial", size=7)
    cond_font_i = Font(name="Arial", italic=True, size=7)
    wrap_align = Alignment(wrap_text=True)

    # Planilla de Cómputo footnote — goes above the standard conditions.
    if _pdf_has_m2_override(data):
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(r, 1).value = _M2_OVERRIDE_FOOTNOTE
        ws.cell(r, 1).font = cond_font_i
        ws.cell(r, 1).alignment = wrap_align
        ws.row_dimensions[r].height = 28
        r += 2

    def _write_cond_line(row, text, font=None):
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws.cell(row, 1)
        cell.value = text
        cell.font = font or cond_font
        cell.alignment = wrap_align

    _write_cond_line(r, "Forma de pago: Contado")
    r += 1
    if co.get("conditions_general"):
        for line in co["conditions_general"].split("\n"):
            if line.strip():
                _write_cond_line(r, line.strip())
                r += 1
    if co.get("conditions_payment"):
        r += 1
        for line in co["conditions_payment"].split("\n"):
            if line.strip():
                _write_cond_line(r, line.strip())
                r += 1

    # Footer — match PDF italic 7pt
    r += 1
    _write_cond_line(r, "No se suben mesadas que no entren en ascensor",
                     Font(name="Arial", italic=True, size=7))
    # Edificio-only: las piezas se dejan en pie de obra (no se colocan)
    r += 1
    _write_cond_line(r, "Las mesadas se dejan en pie de obra",
                     Font(name="Arial", italic=True, size=7))

    wb.save(str(excel_path))
    _force_arial_font(str(excel_path))
    _inject_locale(str(excel_path))


def _generate_pdf(pdf_path: Path, data: dict) -> None:
    """Generate clean PDF using fpdf2 — matches Excel content."""
    FPDF = _make_safe_fpdf()

    client_name = data.get("client_name", "")
    project = data.get("project", "")
    date_str = datetime.now().strftime("%d/%m/%Y")
    delivery = _normalize_delivery(data.get("delivery_days", ""))
    mat_name = data.get("material_name", "")
    mat_m2 = data.get("material_m2", 0)
    mat_price = data.get("material_price_unit", 0)
    currency = data.get("material_currency", "USD")
    discount_pct = data.get("discount_pct", 0)
    sectors = data.get("sectors", [])
    sinks = data.get("sinks", [])
    mo_items = data.get("mo_items", [])
    total_ars = data.get("total_ars", 0)
    total_usd = data.get("total_usd", 0)

    # PR #424 — modo products_only: cotización solo de pileta/bacha
    # como producto suelto, sin material+MO+colocación. El render gate
    # cubre el caso aunque el flag no esté presente — basta con que
    # `material_m2==0 and not material_name` para no emitir bloque
    # material vacío. Eso protege también casos legacy persistidos
    # antes de #424 (regen PDF de quote viejo).
    _is_products_only = data.get("_quote_mode") == "products_only"
    _has_material_block = bool(mat_name) and (mat_m2 or 0) > 0

    co = _load_company_config()

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Top bar
    pdf.set_fill_color(26, 47, 94)
    pdf.rect(0, 0, 210, 4, "F")

    # Logo image
    logo_path = TEMPLATES_DIR / "logo-dangelo.png"
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=10, w=55)
        pdf.set_y(35)
    else:
        pdf.set_y(12)
        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 12, co["name"], new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(0, 4, co["subtitle"], new_x="LMARGIN", new_y="NEXT")

    # Contact
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, co["address"], new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, co["phone"], new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, co["email"], new_x="LMARGIN", new_y="NEXT")

    # Title
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "Presupuesto", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, f"Fecha: {date_str}", new_x="LMARGIN", new_y="NEXT")

    # Client grid
    pdf.ln(3)
    col_w = 95
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w, 5, "Cliente:")
    pdf.cell(col_w, 5, "Forma de pago", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(col_w, 5, client_name)
    pdf.cell(col_w, 5, "Contado", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w, 5, "Proyecto")
    pdf.cell(col_w, 5, "Fecha de entrega", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    # PR #43 — si el proyecto es muy largo (ej: EGEA
    # 'Licitación - Ampliación Unidad Penitenciaria N°12 - Rosario - Edificio B'),
    # la cell desborda y pisa la columna 'Fecha de entrega'. Truncar con
    # ellipsis si no cabe en col_w con un margen de 2mm.
    _proj_disp = project or ""
    try:
        _max_w = col_w - 2
        while _proj_disp and pdf.get_string_width(_proj_disp) > _max_w:
            _proj_disp = _proj_disp[:-1]
        if _proj_disp != (project or ""):
            _proj_disp = _proj_disp[:-1] + "…" if _proj_disp else "…"
    except Exception:
        pass
    pdf.cell(col_w, 5, _proj_disp)
    pdf.cell(col_w, 5, str(delivery), new_x="LMARGIN", new_y="NEXT")

    # Separator
    pdf.ln(2)
    pdf.set_draw_color(26, 26, 26)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    # Table header
    # PR #14 — ampliar columna Descripción (92→112mm) para que entren labels
    # de tipologías edificio largos como "1.45 × 0.50 ME04b-B Mesada recta
    # c/zócalo h:5cm c/frente h:5cm * (×2)". Reduce levemente las columnas
    # de números (~total se mantiene en 190mm).
    w = [112, 18, 30, 30]
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(w[0], 6, "Descripcion")
    pdf.cell(w[1], 6, "Cantidad", align="R")
    pdf.cell(w[2], 6, "Precio unitario", align="R")
    pdf.cell(w[3], 6, "Precio total", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)

    # Material row — Argentine format
    fmt_price = _fmt_usd if currency == "USD" else _fmt_ars
    price_fmt = fmt_price(mat_price)
    total_mat_bruto = round(mat_m2 * mat_price)
    total_mat_fmt = fmt_price(total_mat_bruto)  # Show BRUTO in header row
    if discount_pct:
        total_mat = round(total_mat_bruto * (1 - discount_pct / 100))
    else:
        total_mat = total_mat_bruto

    # Alternating row helper — continuous counter across ALL content rows
    row_n = [0]
    rh = 5  # row height
    total_w = w[0] + w[1] + w[2] + w[3]

    def row_fill():
        """Returns True and sets gray fill for odd rows."""
        fill = row_n[0] % 2 == 1
        if fill:
            pdf.set_fill_color(243, 243, 243)
        return fill

    def row_done():
        row_n[0] += 1

    # PR #424 — gate de TODO el bloque material. En modo products_only
    # (o cualquier data que llegue sin material), NO emitimos el header
    # del material vacío con $0 ni el overlay de descuento del bloque.
    # El descuento, si lo hay, se renderiza como línea aparte después
    # del bloque sinks (más abajo).
    if _has_material_block:
        # Material header row
        f = row_fill()
        pdf.set_font("Helvetica", "B", 9)
        # Thickness: use from breakdown, skip if name already contains Nmm/NMM
        import re as _re
        _thickness = data.get("thickness_mm", 20)
        _mat_display = f"{mat_name} - {_thickness}mm" if not _re.search(r'\d+[Mm][Mm]', mat_name) else mat_name
        pdf.cell(w[0], rh, _mat_display, fill=f)
        pdf.cell(w[1], rh, _fmt_qty(mat_m2), align="R", fill=f)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(w[2], rh, price_fmt, align="R", fill=f)
        pdf.cell(w[3], rh, total_mat_fmt, align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

    # Collect all pieces across sectors for consecutive layout
    all_piece_rows = []  # list of (is_sector_header, display_text)
    for sector in sectors:
        all_piece_rows.append((True, sector.get("label", "")))
        raw_pieces = sector.get("pieces", [])
        grouped, seen = [], {}
        for p in raw_pieces:
            if p in seen:
                seen[p] += 1
            else:
                seen[p] = 1
                grouped.append(p)
        for piece in grouped:
            count = seen[piece]
            display = f"{piece} (×{count})" if count > 1 else piece
            all_piece_rows.append((False, display))

    # Build right-side overlay rows (Descuento + TOTAL)
    # Applies to both USD and ARS — originally gated to USD only which
    # hid the discount line for ARS edificios (bug DINALE 14/04/2026).
    #
    # PR #420 — `TOTAL ARS` se elimina del bloque material por pedido
    # del operador: era un subtotal intermedio ruidoso que confundía
    # al cliente. El único total visible en cotizaciones ARS es el
    # PRESUPUESTO TOTAL al final del documento. La línea `Descuento N%`
    # con monto en negativo se MANTIENE — junto con el grand total
    # final permite reconstruir mentalmente la cuenta (material bruto −
    # descuento + MO = grand total).
    #
    # USD se mantiene porque la convención del operador es distinta:
    # cotizaciones USD muestran el subtotal en USD en el bloque material
    # para que el cliente vea el monto en dólares antes del grand total
    # mixto (USD material + ARS MO).
    # PR #424 — `right_rows` (Descuento + TOTAL USD overlay del bloque
    # material) solo aplican cuando hay bloque material. En modo
    # products_only el descuento se renderiza más abajo, después del
    # bloque de sinks (línea aparte, no overlay).
    right_rows = []
    if _has_material_block:
        if discount_pct:
            desc_amount = round(total_mat_bruto * discount_pct / 100)
            right_rows.append(("I", f"Descuento {discount_pct}%", f"- {fmt_price(desc_amount)}"))
            if currency == "USD":
                right_rows.append(("B", f"TOTAL {currency}", fmt_price(total_mat)))
        elif currency == "USD":
            # Keep legacy behavior for USD without discount — always show TOTAL USD
            # row so the user sees the net material total in the material block.
            right_rows.append(("B", f"TOTAL {currency}", fmt_price(total_mat)))

    # PR #34 — overlay RIGHT COLUMN content (Descuento / TOTAL USD/ARS) en
    # las PRIMERAS filas de piezas, no las últimas. Antes se ponía al final,
    # quedando el TOTAL USD flotando cerca de la sección MANO DE OBRA y
    # confundiendo al cliente. Ahora aparece inmediatamente debajo del
    # subtotal de material — junto a las primeras piezas del despiece.
    piece_indices = [i for i, (is_hdr, _) in enumerate(all_piece_rows) if not is_hdr]
    overlay_map = {}  # piece_index -> right_row_index
    for ri, pi_offset in enumerate(range(0, min(len(right_rows), len(piece_indices)))):
        overlay_map[piece_indices[pi_offset]] = ri

    # Render all rows consecutively
    for idx, (is_hdr, text) in enumerate(all_piece_rows):
        f = row_fill()
        if is_hdr:
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(total_w, rh, text, fill=f, new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(w[0], rh, text, fill=f)
            if idx in overlay_map:
                style, label, value = right_rows[overlay_map[idx]]
                pdf.set_font("Helvetica", style, 8)
                pdf.cell(w[1], rh, "", fill=f)
                pdf.cell(w[2], rh, label, align="R", fill=f)
                pdf.cell(w[3], rh, value, align="R", fill=f)
            else:
                pdf.cell(w[1] + w[2] + w[3], rh, "", fill=f)
            pdf.ln()
        row_done()

    # If fewer pieces than right_rows, add remaining right-side rows
    remaining = len(right_rows) - len([pi for pi in overlay_map])
    if remaining < len(right_rows) and len(piece_indices) < len(right_rows):
        for ri in range(len(right_rows)):
            if ri not in overlay_map.values():
                f = row_fill()
                style, label, value = right_rows[ri]
                pdf.set_font("Helvetica", style, 8)
                pdf.cell(w[0], rh, "", fill=f)
                pdf.cell(w[1], rh, "", fill=f)
                pdf.cell(w[2], rh, label, align="R", fill=f)
                pdf.cell(w[3], rh, value, align="R", fill=f)
                pdf.ln()
                row_done()

    # Spacer row
    pdf.ln(3)

    # Sinks
    for sink in sinks:
        f = row_fill()
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(w[0], rh, sink["name"], fill=f)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(w[1], rh, str(sink["quantity"]), align="R", fill=f)
        pdf.cell(w[2], rh, _fmt_ars(sink['unit_price']), align="R", fill=f)
        pdf.cell(w[3], rh, _fmt_ars(sink['unit_price'] * sink['quantity']), align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

    # PR #424 — Descuento del bloque sinks en modo products_only.
    # En el flujo normal el descuento se renderiza como overlay del
    # bloque material; acá no hay bloque material, así que va como
    # línea propia después de los sinks. Coincide con la convención
    # del PR #420 (descuento siempre en negativo).
    if _is_products_only and discount_pct and sinks:
        sinks_subtotal = sum(s["unit_price"] * s["quantity"] for s in sinks)
        _disc_sinks = round(sinks_subtotal * discount_pct / 100)
        f = row_fill()
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(w[0] + w[1], rh, f"Descuento {discount_pct}%", fill=f)
        pdf.cell(w[2], rh, "", fill=f)
        pdf.cell(w[3], rh, f"- {_fmt_ars(_disc_sinks)}", align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

    # ── SOBRANTE (merma) — línea separada sumada al grand total ──
    # Regla: "Bloque separado e independiente, subtotal propio. Grand total
    # suma principal + sobrante" (rules/calculation-formulas.md).
    sobrante_m2_pdf = data.get("sobrante_m2", 0)
    sobrante_total_pdf = data.get("sobrante_total", 0)
    if sobrante_m2_pdf and sobrante_total_pdf:
        pdf.ln(2)
        f = row_fill()
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(total_w, rh, "SOBRANTE (MERMA)", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()
        f = row_fill()
        pdf.set_font("Helvetica", "", 9)
        # Precio unit depende de la moneda del material
        _sob_unit = mat_price  # mismo precio unitario que el material
        _sob_fmt = _fmt_usd if currency == "USD" else _fmt_ars
        pdf.cell(w[0], rh, "Sobrante disponible", fill=f)
        pdf.cell(w[1], rh, f"{sobrante_m2_pdf:.2f}".replace(".", ","), align="R", fill=f)
        pdf.cell(w[2], rh, _sob_fmt(_sob_unit), align="R", fill=f)
        pdf.cell(w[3], rh, _sob_fmt(sobrante_total_pdf), align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

    # PR #424 — Bloque MO solo si hay items. En modo products_only
    # mo_items=[] y no hay header "MANO DE OBRA" suelto. Esto también
    # cubre cualquier flujo legacy donde se llegue acá sin MO (raro
    # pero no rompible).
    if mo_items:
        # Spacer
        pdf.ln(2)

        # MO header
        f = row_fill()
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(total_w, rh, "MANO DE OBRA", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

        # MO items
        for mo in mo_items:
            f = row_fill()
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(w[0], rh, mo["description"], fill=f)
            pdf.cell(w[1], rh, _fmt_qty(mo["quantity"]), align="R", fill=f)
            pdf.cell(w[2], rh, _fmt_ars(mo['unit_price']), align="R", fill=f)
            pdf.cell(w[3], rh, _fmt_ars(round(mo['unit_price'] * mo['quantity'])), align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
            row_done()

    # MO commercial discount line (edificio "% sobre MO") — before Total PESOS
    _mo_disc_pct = data.get("mo_discount_pct", 0)
    _mo_disc_amt = data.get("mo_discount_amount", 0)
    if _mo_disc_pct and _mo_disc_amt:
        f = row_fill()
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(w[0] + w[1], rh, f"Descuento {_mo_disc_pct}% sobre MO (excluye flete)", fill=f)
        pdf.cell(w[2], rh, "", fill=f)
        pdf.cell(w[3], rh, f"- {_fmt_ars(_mo_disc_amt)}", align="R", fill=f, new_x="LMARGIN", new_y="NEXT")
        row_done()

    # Total PESOS
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(w[0] + w[1], 5, "")
    pdf.cell(w[2], 5, "Total PESOS", align="R")
    pdf.cell(w[3], 5, _fmt_ars(total_ars), align="R", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(5)

    # Grand total box
    grand = _format_grand_total(total_ars, total_usd, currency)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_draw_color(26, 26, 26)
    y_box = pdf.get_y()
    pdf.rect(10, y_box, 190, 8)
    pdf.set_xy(10, y_box + 1.5)
    pdf.cell(190, 5, grand, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)

    # Planilla de Cómputo footnote (above legends, italic small)
    if _pdf_has_m2_override(data):
        pdf.set_font("Helvetica", "I", 7)
        pdf.multi_cell(
            0, 3.5, _M2_OVERRIDE_FOOTNOTE,
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(0.5)

    # Footer note
    pdf.set_font("Helvetica", "BI", 8)
    pdf.cell(0, 4, "No se suben mesadas que no entren en ascensor", new_x="LMARGIN", new_y="NEXT")
    # Edificio-only extra legend
    if data.get("is_edificio"):
        pdf.cell(0, 4, "Las mesadas se dejan en pie de obra", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)

    # Conditions (from config.json) — use multi_cell for word wrap
    usable_w = 190  # A4 width (210) minus margins (10+10)
    if co["conditions_general"]:
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(0, 3, "CONDICIONES", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 7)
        for line in co["conditions_general"].split("\n"):
            if line.strip():
                pdf.multi_cell(usable_w, 3, line.strip(), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    if co["conditions_payment"]:
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(0, 3, "FORMAS DE PAGO", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 7)
        for line in co["conditions_payment"].split("\n"):
            if line.strip():
                pdf.multi_cell(usable_w, 3, line.strip(), new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(pdf_path))


def _generate_excel(output_path: Path, data: dict) -> None:
    """Generate Excel from template — only replace values, keep all formatting."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    TEMPLATE = TEMPLATES_DIR / "excel" / "quote-template.xlsx"
    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb.active

    # Unmerge all cells first to avoid MergedCell write errors
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    # Reset any hidden rows + row heights in the dynamic band so the output
    # never inherits stale 0-height / hidden rows from the source template.
    # openpyxl stores row settings in ws.row_dimensions[idx].
    for _rd_idx in range(20, 200):
        _rd = ws.row_dimensions.get(_rd_idx)
        if _rd is not None:
            _rd.hidden = False
            _rd.height = None  # Let Excel auto-size
    # Also clear any workbook-level conditional formatting that applies to the
    # dynamic band (template zebra rules fire on rows we write to).
    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass

    client_name = data.get("client_name", "")
    project = data.get("project", "")
    date_str = datetime.now().strftime("%d/%m/%Y")
    delivery = _normalize_delivery(data.get("delivery_days", ""))
    mat_name = data.get("material_name", "")
    mat_m2 = data.get("material_m2", 0)
    mat_price = data.get("material_price_unit", 0)
    currency = data.get("material_currency", "USD")
    discount_pct = data.get("discount_pct", 0)
    sectors = data.get("sectors", [])
    mo_items = data.get("mo_items", [])
    sinks = data.get("sinks", [])
    total_ars = data.get("total_ars", 0)
    total_usd = data.get("total_usd", 0)

    # PR #424 — modo products_only (mismo gate que el PDF). Si no hay
    # material real, no escribimos nada en row 22 (mat header) ni en
    # el overlay de descuento del bloque material. Excel y PDF
    # mantienen consistencia visual.
    _is_products_only = data.get("_quote_mode") == "products_only"
    _has_material_block = bool(mat_name) and (mat_m2 or 0) > 0

    bold = Font(name="Arial", bold=True, size=10)
    normal = Font(name="Arial", bold=False, size=10)
    small = Font(name="Arial", bold=False, size=9)
    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    # ── Zebra striping (matches PDF F3F3F3 on odd rows) ────────────────────
    zebra_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    _row_n = [0]

    def _apply_zebra(sheet_row: int, cols: int = 6) -> None:
        fill = zebra_fill if _row_n[0] % 2 == 1 else no_fill
        for c in range(1, cols + 1):
            ws.cell(sheet_row, c).fill = fill

    def _zebra_done() -> None:
        _row_n[0] += 1

    ars_fmt = '"$"#,##0.00'
    # PR #37 — formato de cantidades preserva hasta 4 decimales sin
    # trailing zeros. Antes '#,##0.00' redondeaba a 2 decimales en el
    # display — cuando el operador editaba 4.295 en Google Sheets/Excel,
    # veía 4.30 (aunque el valor interno era correcto). Ahora:
    #  4      → '4'
    #  4.5    → '4,5'
    #  4.30   → '4,3'
    #  4.295  → '4,295'
    #  4.2955 → '4,2955'
    qty_fmt = 'General'

    # Header — replace values, keep template formatting
    ws["A13"].value = f"Fecha: {date_str}"
    ws["A16"].value = client_name
    ws["A18"].value = project
    ws["C19"].value = delivery

    # Material row 22
    total_mat = round(mat_m2 * mat_price)
    total_mat_net = total_mat
    if discount_pct:
        total_mat_net = total_mat - round(total_mat * discount_pct / 100)

    # PR #424 — Material header row solo si hay material. En modo
    # products_only dejamos row 22 vacío (sin escribir mat_name=""
    # ni cantidades $0,00 ruidosas). El template ya tiene el grid
    # dibujado, así que la fila aparece en blanco — limpia.
    if _has_material_block:
        import re as _re_xl
        _thickness_xl = data.get("thickness_mm", 20)
        ws["A22"].value = f"{mat_name} - {_thickness_xl}mm" if not _re_xl.search(r'\d+[Mm][Mm]', mat_name) else mat_name
        ws["D22"].value = mat_m2
        ws["D22"].number_format = qty_fmt
        if currency == "USD":
            ws["E22"].value = _fmt_usd(mat_price)
            ws["F22"].value = _fmt_usd(total_mat)  # Bruto (before discount)
        else:
            ws["E22"].value = mat_price
            ws["E22"].number_format = ars_fmt
            ws["F22"].value = total_mat  # Bruto (before discount)
            ws["F22"].number_format = ars_fmt
        # Zebra row 0 (material header) — no fill
        _apply_zebra(22)
        _zebra_done()

    # ── ONLY replace .value — NEVER touch .font, .fill, .alignment, .border ──
    # The template has all formatting correct. We just swap the data.

    # Pieces (group duplicates) — compute up front to decide how many rows
    # to insert BEFORE the MO block. The template only has 4 piece slots
    # (rows 23-26); with more pieces MO would overlap despiece.
    raw_all = []
    for sector in sectors:
        for piece in sector.get("pieces", []):
            raw_all.append(piece)
    all_pieces = []
    seen_xl = {}
    for p in raw_all:
        if p in seen_xl:
            seen_xl[p] += 1
        else:
            seen_xl[p] = 1
            all_pieces.append(p)
    all_pieces = [f"{p} (×{seen_xl[p]})" if seen_xl[p] > 1 else p for p in all_pieces]

    TEMPLATE_PIECE_SLOTS = 4
    MO_HEADER_ROW_BASE = 27
    extra_pieces = max(0, len(all_pieces) - TEMPLATE_PIECE_SLOTS)
    # Reservar filas para piletas: 1 header + N rows (si hay piletas).
    sinks_block_size = (1 + len(sinks)) if sinks else 0
    extra_for_sinks = sinks_block_size
    # PR #424 — reserva 1 fila extra para el descuento de sinks en
    # modo products_only. Se inserta después del último sink, antes
    # de SOBRANTE / MO header.
    extra_for_products_disc = (
        1 if (_is_products_only and discount_pct and sinks) else 0
    )
    extra_for_sinks += extra_for_products_disc
    _sob_m2_pre = data.get("sobrante_m2", 0)
    _sob_total_pre = data.get("sobrante_total", 0)
    extra_for_sobrante_pre = 2 if (_sob_m2_pre and _sob_total_pre) else 0
    total_extra = extra_pieces + extra_for_sinks + extra_for_sobrante_pre
    if total_extra > 0:
        ws.insert_rows(MO_HEADER_ROW_BASE, total_extra)

    # Clear values AND fills in dynamic rows.
    max_clear = 35 + total_extra + max(0, len(mo_items) - 4) + 5
    for row in range(23, max_clear + 1):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = no_fill

    # Build right-side overlay (Descuento + TOTAL)
    # Applies to both USD and ARS — originally gated to USD only which hid
    # the discount line for ARS edificios (bug DINALE 14/04/2026).
    #
    # PR #420 — `TOTAL ARS` se elimina del bloque material por pedido
    # del operador (mismo cambio que el PDF). USD se mantiene porque
    # aclara al cliente el monto en dólares antes del grand total mixto.
    # Ver document_tool.py:1428 (PDF) para el racional completo.
    # PR #424 — `right_rows_xl` (Descuento + TOTAL USD overlay del
    # bloque material) solo si hay bloque material. Mismo criterio
    # que el PDF: en modo products_only el descuento se renderiza
    # como línea aparte después del bloque sinks.
    italic_sm = Font(name="Arial", italic=True, size=9)
    _xl_fmt = _fmt_usd if currency == "USD" else _fmt_ars
    right_rows_xl = []
    if _has_material_block:
        if discount_pct:
            desc_amount = round(total_mat * discount_pct / 100)
            right_rows_xl.append(("I", f"Descuento {discount_pct}%", f"- {_xl_fmt(desc_amount)}"))
            if currency == "USD":
                right_rows_xl.append(("B", f"TOTAL {currency}", _xl_fmt(total_mat_net)))
        elif currency == "USD":
            right_rows_xl.append(("B", f"TOTAL {currency}", _xl_fmt(total_mat_net)))

    # PR #34 — overlay en las PRIMERAS piezas (no las últimas), para que
    # TOTAL USD/ARS aparezca inmediatamente debajo del subtotal del
    # material y no floating al final del despiece.
    r = 23
    for i, piece in enumerate(all_pieces):
        ws.cell(r, 1).value = piece
        ri = i  # ri direct index — primeras N filas obtienen right_row
        if 0 <= ri < len(right_rows_xl):
            style, label, value = right_rows_xl[ri]
            ws.cell(r, 5).value = label
            ws.cell(r, 5).font = italic_sm if style == "I" else bold
            ws.cell(r, 5).alignment = right_align
            ws.cell(r, 6).value = value
            ws.cell(r, 6).font = italic_sm if style == "I" else bold
            ws.cell(r, 6).alignment = right_align
        _apply_zebra(r)
        _zebra_done()
        r += 1

    # If fewer pieces than right_rows, add remaining
    if len(all_pieces) < len(right_rows_xl):
        for ri in range(len(all_pieces), len(right_rows_xl)):
            style, label, value = right_rows_xl[ri]
            ws.cell(r, 5).value = label
            ws.cell(r, 5).font = italic_sm if style == "I" else bold
            ws.cell(r, 5).alignment = right_align
            ws.cell(r, 6).value = value
            ws.cell(r, 6).font = italic_sm if style == "I" else bold
            ws.cell(r, 6).alignment = right_align
            _apply_zebra(r)
            _zebra_done()
            r += 1

    # ── PILETAS / SINKS — insertar entre piezas y MO header (igual que PDF) ─
    sinks_header_row = MO_HEADER_ROW_BASE + extra_pieces
    if sinks:
        ws.cell(sinks_header_row, 1).value = "PILETAS"
        ws.cell(sinks_header_row, 1).font = bold
        _apply_zebra(sinks_header_row)
        _zebra_done()
        for i, sink in enumerate(sinks):
            srow = sinks_header_row + 1 + i
            ws.cell(srow, 1).value = sink.get("name", "")
            ws.cell(srow, 4).value = sink.get("quantity", 1)
            ws.cell(srow, 4).number_format = qty_fmt
            ws.cell(srow, 5).value = sink.get("unit_price", 0)
            ws.cell(srow, 5).number_format = ars_fmt
            ws.cell(srow, 6).value = f"=D{srow}*E{srow}"
            ws.cell(srow, 6).number_format = ars_fmt
            _apply_zebra(srow)
            _zebra_done()

        # PR #424 — descuento de sinks en modo products_only. Igual
        # criterio que PDF: una fila al final del bloque sinks con
        # `Descuento N%` y monto en negativo. NO se aplica al modo
        # normal (donde el descuento ya va en el overlay del bloque
        # material).
        if _is_products_only and discount_pct:
            sinks_subtotal = sum(
                (s.get("unit_price", 0) or 0) * (s.get("quantity", 0) or 0)
                for s in sinks
            )
            _disc_sinks = round(sinks_subtotal * discount_pct / 100)
            disc_row = sinks_header_row + 1 + len(sinks)
            ws.cell(disc_row, 1).value = f"Descuento {discount_pct}%"
            ws.cell(disc_row, 1).font = italic_sm
            ws.cell(disc_row, 6).value = -_disc_sinks
            ws.cell(disc_row, 6).number_format = ars_fmt
            ws.cell(disc_row, 6).font = italic_sm
            _apply_zebra(disc_row)
            _zebra_done()

    # ── SOBRANTE (merma) — línea separada al grand total ──
    sobrante_m2_xl = data.get("sobrante_m2", 0)
    sobrante_total_xl = data.get("sobrante_total", 0)
    extra_for_sobrante = 2 if (sobrante_m2_xl and sobrante_total_xl) else 0
    if extra_for_sobrante:
        sob_header_row = sinks_header_row + extra_for_sinks
        ws.cell(sob_header_row, 1).value = "SOBRANTE (MERMA)"
        ws.cell(sob_header_row, 1).font = bold
        _apply_zebra(sob_header_row)
        _zebra_done()
        sob_row = sob_header_row + 1
        ws.cell(sob_row, 1).value = "Sobrante disponible"
        ws.cell(sob_row, 4).value = sobrante_m2_xl
        ws.cell(sob_row, 4).number_format = qty_fmt
        ws.cell(sob_row, 5).value = mat_price  # mismo precio que material
        if currency == "USD":
            ws.cell(sob_row, 5).value = _fmt_usd(mat_price)
            ws.cell(sob_row, 6).value = _fmt_usd(sobrante_total_xl)
        else:
            ws.cell(sob_row, 5).number_format = ars_fmt
            ws.cell(sob_row, 6).value = sobrante_total_xl
            ws.cell(sob_row, 6).number_format = ars_fmt
        _apply_zebra(sob_row)
        _zebra_done()

    # MO items — header shifted by extra_pieces + sinks + sobrante.
    MO_HEADER_ROW = MO_HEADER_ROW_BASE + extra_pieces + extra_for_sinks + extra_for_sobrante
    MO_START_ROW = MO_HEADER_ROW + 1
    TEMPLATE_MO_SLOTS = 4
    extra_mo = max(0, len(mo_items) - TEMPLATE_MO_SLOTS)
    if extra_mo > 0:
        ws.insert_rows(MO_START_ROW + TEMPLATE_MO_SLOTS, extra_mo)

    # PR #424 — MO header solo si hay items. En modo products_only
    # mo_items=[] y la fila 27 queda vacía (limpia, sin "MANO DE
    # OBRA" suelto sobre nada).
    if mo_items:
        ws.cell(MO_HEADER_ROW, 1).value = "MANO DE OBRA"
        _apply_zebra(MO_HEADER_ROW)
        _zebra_done()
        for i, mo in enumerate(mo_items):
            row = MO_START_ROW + i
            ws.cell(row, 1).value = mo["description"]
            ws.cell(row, 4).value = mo["quantity"]
            ws.cell(row, 4).number_format = qty_fmt
            ws.cell(row, 5).value = mo["unit_price"]
            ws.cell(row, 5).number_format = ars_fmt
            ws.cell(row, 6).value = f"=D{row}*E{row}"
            ws.cell(row, 6).number_format = ars_fmt
            _apply_zebra(row)
            _zebra_done()

    # MO commercial discount line (edificio "% sobre MO") — before Total PESOS
    # PR #424 — si no hay mo_items, `mo_end_row` apuntaría al header
    # del bloque MO (que tampoco se rendereó). Ajustamos para que
    # `Total PESOS` quede al inicio del bloque MO reservado.
    if mo_items:
        mo_end_row = MO_START_ROW + len(mo_items) - 1
    else:
        # Sin MO: dejamos el `Total PESOS` justo donde habría ido el
        # header MO. Saltamos ese row vacío para que no quede
        # arrastrando demasiado espacio en blanco entre sinks y total.
        mo_end_row = MO_HEADER_ROW - 1
    _mo_disc_pct = data.get("mo_discount_pct", 0)
    _mo_disc_amt = data.get("mo_discount_amount", 0)
    if _mo_disc_pct and _mo_disc_amt:
        disc_row = mo_end_row + 1
        ws.insert_rows(disc_row)
        ws.cell(disc_row, 1).value = f"Descuento {_mo_disc_pct}% sobre MO (excluye flete)"
        ws.cell(disc_row, 1).font = italic_sm
        ws.cell(disc_row, 6).value = -_mo_disc_amt
        ws.cell(disc_row, 6).number_format = ars_fmt
        ws.cell(disc_row, 6).font = italic_sm
        _apply_zebra(disc_row)
        _zebra_done()
        mo_end_row = disc_row

    # Total PESOS — after all MO items (and discount line if present)
    total_pesos_row = mo_end_row + 1
    ws.cell(total_pesos_row, 5).value = "Total PESOS"
    ws.cell(total_pesos_row, 5).font = bold
    ws.cell(total_pesos_row, 6).value = total_ars
    ws.cell(total_pesos_row, 6).number_format = ars_fmt
    ws.cell(total_pesos_row, 6).font = bold
    _apply_zebra(total_pesos_row)
    _zebra_done()

    # Grand total — 2 rows after total pesos
    grand_row = total_pesos_row + 2
    grand = _format_grand_total(total_ars, total_usd, currency)
    # Clear template remnants around grand total, and wipe a bigger band
    # below to avoid stale example-piece rows surviving from the source
    # template (ZOC rows etc.) when we inserted many pieces.
    for clear_row in range(total_pesos_row + 1, grand_row + 20):
        for clear_col in range(1, 7):
            cell = ws.cell(clear_row, clear_col)
            cell.value = None
            cell.border = Border()  # Remove stale template borders
    ws.cell(grand_row, 1).value = grand
    ws.cell(grand_row, 1).alignment = center_align
    ws.merge_cells(f"A{grand_row}:F{grand_row}")
    # Apply box border + center align + no fill to grand total row
    for col in range(1, 7):
        ws.cell(grand_row, col).border = box
        ws.cell(grand_row, col).font = bold
        ws.cell(grand_row, col).alignment = center_align
        ws.cell(grand_row, col).fill = no_fill

    # Conditions block (CONDICIONES / FORMAS DE PAGO) + footer legend(s).
    # Mirrors the PDF layout so Excel is not a stripped-down version.
    co = _load_company_config()
    cond_font_b = Font(name="Arial", bold=True, size=7)
    cond_font = Font(name="Arial", size=7)
    cond_font_i = Font(name="Arial", italic=True, size=7)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    cr = grand_row + 2

    # Planilla de Cómputo footnote (above conditions) — only if any piece
    # used m2_override.
    if _pdf_has_m2_override(data):
        ws.cell(cr, 1).value = _M2_OVERRIDE_FOOTNOTE
        ws.cell(cr, 1).font = cond_font_i
        ws.cell(cr, 1).alignment = wrap_align
        ws.merge_cells(f"A{cr}:F{cr}")
        ws.row_dimensions[cr].height = 28
        cr += 2

    if co.get("conditions_general"):
        ws.cell(cr, 1).value = "CONDICIONES"
        ws.cell(cr, 1).font = cond_font_b
        ws.merge_cells(f"A{cr}:F{cr}")
        cr += 1
        for line in co["conditions_general"].split("\n"):
            if line.strip():
                ws.cell(cr, 1).value = line.strip()
                ws.cell(cr, 1).font = cond_font
                ws.cell(cr, 1).alignment = wrap_align
                ws.merge_cells(f"A{cr}:F{cr}")
                cr += 1
        cr += 1

    if co.get("conditions_payment"):
        ws.cell(cr, 1).value = "FORMAS DE PAGO"
        ws.cell(cr, 1).font = cond_font_b
        ws.merge_cells(f"A{cr}:F{cr}")
        cr += 1
        for line in co["conditions_payment"].split("\n"):
            if line.strip():
                ws.cell(cr, 1).value = line.strip()
                ws.cell(cr, 1).font = cond_font
                ws.cell(cr, 1).alignment = wrap_align
                ws.merge_cells(f"A{cr}:F{cr}")
                cr += 1
        cr += 1

    # Footer legend — always show "No se suben mesadas…"; edificio adds the
    # "Las mesadas se dejan en pie de obra" line.
    legend_font = Font(name="Arial", italic=True, bold=True, size=8)
    center_legend = Alignment(horizontal="center", vertical="center")
    ws.cell(cr, 1).value = "No se suben mesadas que no entren en ascensor"
    ws.cell(cr, 1).font = legend_font
    ws.cell(cr, 1).alignment = center_legend
    ws.merge_cells(f"A{cr}:F{cr}")
    cr += 1
    if data.get("is_edificio"):
        ws.cell(cr, 1).value = "Las mesadas se dejan en pie de obra"
        ws.cell(cr, 1).font = legend_font
        ws.cell(cr, 1).alignment = center_legend
        ws.merge_cells(f"A{cr}:F{cr}")

    wb.save(str(output_path))
    _force_arial_font(str(output_path))
    _inject_locale(str(output_path))


def _force_arial_font(xlsx_path: str) -> None:
    """Replace Calibri → Arial in the saved xlsx file.

    PR #35 — Google Sheets doesn't have Calibri installed. When users open
    the xlsx from Drive, fonts fall back to a system font rendered larger
    at the same point size. Arial exists on both Excel and Sheets → same
    visual size in both. Post-processes styles.xml + sharedStrings if any.
    """
    import zipfile
    import shutil
    from pathlib import Path as _P
    _p = _P(xlsx_path)
    if not _p.exists():
        return
    _tmp = _p.with_suffix(".arial.tmp")
    try:
        with zipfile.ZipFile(str(_p), "r") as zin:
            with zipfile.ZipFile(str(_tmp), "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.endswith(".xml"):
                        try:
                            text = data.decode("utf-8")
                            text = text.replace('name="Calibri"', 'name="Arial"')
                            text = text.replace('val="Calibri"', 'val="Arial"')
                            data = text.encode("utf-8")
                        except UnicodeDecodeError:
                            pass
                    zout.writestr(item, data)
        shutil.move(str(_tmp), str(_p))
    except Exception as e:
        logging.warning(f"[force_arial] failed for {_p.name}: {e}")
        try:
            _tmp.unlink()
        except Exception:
            pass


def _inject_locale(xlsx_path: str):
    """Inject es_AR locale into xlsx for Google Sheets.

    Modifies xl/workbook.xml to add spreadsheetLocale="es_AR" to <calcPr>,
    and adds SpreadsheetLocale custom property to docProps/custom.xml.
    Both methods together ensure Google Sheets picks up the locale.
    """
    import zipfile, os, re

    custom_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
            xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2" name="SpreadsheetLocale">
    <vt:lpwstr>es_AR</vt:lpwstr>
  </property>
</Properties>'''

    try:
        tmp = xlsx_path + ".tmp"
        with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == 'xl/workbook.xml':
                    text = data.decode('utf-8')
                    if '<calcPr' in text and 'spreadsheetLocale' not in text:
                        text = re.sub(
                            r'<calcPr\b',
                            '<calcPr spreadsheetLocale="es_AR"',
                            text,
                        )
                    elif '<calcPr' not in text:
                        text = text.replace(
                            '</workbook>',
                            '<calcPr spreadsheetLocale="es_AR" calcId="191029"/></workbook>',
                        )
                    data = text.encode('utf-8')

                elif item.filename == '[Content_Types].xml':
                    text = data.decode('utf-8')
                    if 'custom-properties' not in text:
                        text = text.replace(
                            '</Types>',
                            '<Override PartName="/docProps/custom.xml" ContentType="application/vnd.openxmlformats-officedocument.custom-properties+xml"/></Types>',
                        )
                    data = text.encode('utf-8')

                elif item.filename == '_rels/.rels':
                    text = data.decode('utf-8')
                    if 'custom-properties' not in text:
                        text = text.replace(
                            '</Relationships>',
                            '<Relationship Id="rIdCustom" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/custom-properties" Target="docProps/custom.xml"/></Relationships>',
                        )
                    data = text.encode('utf-8')

                zout.writestr(item, data)
            zout.writestr('docProps/custom.xml', custom_xml)
        os.replace(tmp, xlsx_path)
    except Exception as e:
        import logging
        logging.warning(f"Could not inject locale into xlsx: {e}")
        try:
            os.remove(tmp)
        except OSError:
            pass


def _format_grand_total(total_ars: float, total_usd: float, currency: str) -> str:
    if currency == "USD" and total_usd:
        return f"PRESUPUESTO TOTAL: {_fmt_ars(total_ars)} mano de obra + {_fmt_usd(total_usd)} material"
    return f"PRESUPUESTO TOTAL: {_fmt_ars(total_ars)}"


def generate_comparison_pdf(pdf_path: Path, client_name: str, project: str, quotes_data: list[dict]) -> None:
    """Generate a landscape comparison PDF for multiple material variants."""
    FPDF = _make_safe_fpdf()

    co = _load_company_config()
    date_str = datetime.now().strftime("%d/%m/%Y")
    n = len(quotes_data)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Top bar
    pdf.set_fill_color(26, 47, 94)
    pdf.rect(0, 0, 297, 4, "F")

    # Logo
    logo_path = TEMPLATES_DIR / "logo-dangelo.png"
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=10, w=55)
        pdf.set_y(35)
    else:
        pdf.set_y(12)
        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 12, co["name"], new_x="LMARGIN", new_y="NEXT")

    # Contact
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, f'{co["address"]}  |  {co["phone"]}  |  {co["email"]}', new_x="LMARGIN", new_y="NEXT")

    # Title
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Comparativo de Presupuestos", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Fecha: {date_str}  |  Cliente: {client_name}  |  Proyecto: {project}", new_x="LMARGIN", new_y="NEXT")

    # Separator
    pdf.ln(3)
    pdf.set_draw_color(26, 26, 26)
    pdf.line(10, pdf.get_y(), 287, pdf.get_y())
    pdf.ln(4)

    # Table layout
    label_w = 55
    usable = 277 - label_w  # 297 - 2*10 margins - label
    col_w = usable / n

    # Extract data per variant
    def _mo_total(bd: dict) -> float:
        return sum(m.get("quantity", 0) * m.get("unit_price", 0) for m in bd.get("mo_items", []))

    variants = []
    for bd in quotes_data:
        merma = bd.get("merma", {})
        currency = bd.get("material_currency", "USD")
        fmt_price = _fmt_usd if currency == "USD" else _fmt_ars
        variants.append({
            "material": bd.get("_material") or bd.get("material_name", "—"),
            "currency": currency,
            "price_m2": fmt_price(bd.get("material_price_unit", 0)),
            "m2": _fmt_qty(bd.get("material_m2", 0)) + " m\u00b2",
            "total_mat": fmt_price(bd.get("material_total", 0)),
            "discount": f"{bd.get('discount_pct', 0)}%" if bd.get("discount_pct") else "No aplica",
            "mo_total": _fmt_ars(_mo_total(bd)),
            "merma": merma.get("motivo", "No aplica") if merma else "No aplica",
            "delivery": _normalize_delivery(bd.get("delivery_days", "")),
            "total_ars": bd.get("_total_ars") or bd.get("total_ars", 0),
            "total_usd": bd.get("_total_usd") or bd.get("total_usd", 0),
        })

    # Find cheapest by total_usd (or total_ars if no USD)
    best_idx = 0
    best_val = float("inf")
    for i, v in enumerate(variants):
        val = v["total_usd"] if v["total_usd"] else v["total_ars"]
        if val and val < best_val:
            best_val = val
            best_idx = i

    # Header row — material names
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(26, 47, 94)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(label_w, 8, "", fill=True)
    for i, v in enumerate(variants):
        pdf.cell(col_w, 8, v["material"][:30], align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    # Rows definition
    rows = [
        ("Precio / m\u00b2", "price_m2"),
        ("Superficie", "m2"),
        ("Total Material", "total_mat"),
        ("Descuento", "discount"),
        ("Mano de Obra", "mo_total"),
        ("Merma", "merma"),
        ("Plazo de entrega", "delivery"),
    ]

    row_counter = 0
    for label, key in rows:
        is_odd = row_counter % 2 == 0
        if is_odd:
            pdf.set_fill_color(242, 242, 242)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(label_w, 7, label, fill=True)
        pdf.set_font("Helvetica", "", 9)
        for v in variants:
            pdf.cell(col_w, 7, str(v[key])[:35], align="C", fill=True)
        pdf.ln()
        row_counter += 1

    # Total rows — highlighted
    pdf.ln(2)
    pdf.set_draw_color(26, 47, 94)
    pdf.line(10, pdf.get_y(), 287, pdf.get_y())
    pdf.ln(2)

    for total_label, total_key, fmt_fn in [
        ("TOTAL ARS", "total_ars", _fmt_ars),
        ("TOTAL USD", "total_usd", _fmt_usd),
    ]:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(label_w, 9, total_label)
        for i, v in enumerate(variants):
            val = v[total_key]
            text = fmt_fn(val) if val else "—"
            if i == best_idx:
                pdf.set_text_color(0, 128, 0)
            pdf.cell(col_w, 9, text, align="C")
            pdf.set_text_color(0, 0, 0)
        pdf.ln()

    # Best option indicator
    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(0, 128, 0)
    best_name = variants[best_idx]["material"]
    pdf.cell(0, 5, f"* Opcion mas economica: {best_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    # Footer conditions (from config.json)
    pdf.ln(6)
    pdf.set_font("Helvetica", "", 7)
    if co["conditions_general"]:
        for line in co["conditions_general"].split("\n"):
            if line.strip():
                pdf.cell(0, 3, line.strip(), new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(pdf_path))
